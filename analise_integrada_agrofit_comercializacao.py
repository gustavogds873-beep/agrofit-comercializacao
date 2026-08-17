#!/usr/bin/env python
# -*- coding: utf-8 -*-
# =============================================================================
#  ANALISE INTEGRADA AGROFIT x COMERCIALIZACAO DE AGROTOXICOS
# =============================================================================
#
#  Script completo (Windows / Linux / macOS) que cruza a base de produtos
#  formulados do Agrofit com as bases anuais de comercializacao de agrotoxicos,
#  gerando uma estrutura analitica com granularidades separadas para evitar
#  duplicacao/multiplicacao artificial dos volumes comercializados.
#
#  Abas geradas:
#    00_Resumo                         - indicadores, qualidade do match, rankings
#    00_Legenda                        - dicionario, premissas PF vs IA
#    01_Resumo_Agrofit                 - dimensao produto (cadastro Agrofit)
#    02_Comercializacao_Produtos       - fato PF: 1 linha por Registro x Ano
#    03_Evolucao_Produtos              - wide: produto x anos (volume PF)
#    04_Comercializacao_IA             - fato IA: Registro x Ano x IA
#    05_Vendas_IA                      - agregado: IA x Ano (volumes ponderados)
#    06_Evolucao_IA                    - wide: IA x anos (volume ponderado)
#    07_Comerc_Regional_Produtos       - fato PF: Registro x Ano x UF
#    08_Comerc_Regional_IA             - fato IA: IA x Ano x UF
#    09_Agrofit_Sem_Vendas             - Agrofit sem comercializacao no periodo
#    10_Vendas_Sem_Agrofit             - comercializacao sem match no Agrofit
#
#  Premissas multi-IA (Tabela Bruta):
#    - O mesmo produto formulado aparece em N linhas (uma por IA), repetindo
#      os volumes de PF. O script trata:
#        * perspectiva PRODUTO -> first/max nos volumes PF (NAO soma)
#        * perspectiva IA      -> mantem cada linha e aplica PONDERACAO_IA
#
#  Volumes:
#    - ORIGINAIS (auditoria): VENDAS_TOTAIS, VENDAS_CLIENTE, VENDAS_INDUSTRIA,
#      PRODUCAO_NACIONAL, IMPORTACAO, EXPORTACAO, ESTOQUE_* — valores da planilha,
#      sem ponderacao pelo teor de IA. Nunca sao sobrescritos.
#    - PONDERADOS (analise de IA): VENDA_PONDERADA_IA = volume original x
#      PONDERACAO_IA — quantidade efetiva de ingrediente ativo.
#
#  Uso:
#      python analise_integrada_agrofit_comercializacao.py
#
#  Pastas (criadas automaticamente):
#      dados/  -> agrofitprodutosformulados.csv + arquivos anuais (.xlsx/.csv)
#      saida/  -> Relatorio_Comercializacao_Agrofit.xlsx
#      logs/   -> analise_comercializacao.log
#
#  Variaveis de ambiente opcionais:
#      AGROFIT_DADOS_DIR, AGROFIT_SAIDA_DIR, AGROFIT_LOGS_DIR, AGROFIT_NO_INPUT
# =============================================================================

from __future__ import annotations

import os
import sys
import re
import time
import logging
import unicodedata
import warnings
import traceback
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Dict, Tuple, Any

# -----------------------------------------------------------------------
# Dependencias
# -----------------------------------------------------------------------
def _ensure_packages() -> None:
    required = {"pandas": "pandas", "numpy": "numpy", "openpyxl": "openpyxl"}
    missing = []
    for module_name, pip_name in required.items():
        try:
            __import__(module_name)
        except ImportError:
            missing.append(pip_name)
    if missing:
        print(f"Instalando pacotes faltantes: {', '.join(missing)} ...")
        import subprocess
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "-q"] + missing
        )
        print("Pacotes instalados.\n")


_ensure_packages()

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

warnings.filterwarnings("ignore", category=UserWarning)
warnings.filterwarnings("ignore", category=FutureWarning)
pd.set_option("mode.chained_assignment", None)

# =============================================================================
# CONFIGURACAO
# =============================================================================

SCRIPT_DIR = Path(__file__).resolve().parent
DADOS_DIR = Path(os.environ["AGROFIT_DADOS_DIR"]) if os.environ.get("AGROFIT_DADOS_DIR") else SCRIPT_DIR / "dados"
SAIDA_DIR = Path(os.environ["AGROFIT_SAIDA_DIR"]) if os.environ.get("AGROFIT_SAIDA_DIR") else SCRIPT_DIR / "saida"
LOGS_DIR = Path(os.environ["AGROFIT_LOGS_DIR"]) if os.environ.get("AGROFIT_LOGS_DIR") else SCRIPT_DIR / "logs"

ARQUIVO_AGROFIT_PADRAO = "agrofitprodutosformulados.csv"
ARQUIVO_SAIDA = "Relatorio_Comercializacao_Agrofit.xlsx"
ARQUIVO_LOG = "analise_comercializacao.log"
PISTAS_NOME_AGROFIT = ["agrofit"]

UFS = [
    "RO", "AC", "AM", "RR", "PA", "AP", "TO", "MA", "PI", "CE", "RN", "PB",
    "PE", "AL", "SE", "BA", "MG", "ES", "RJ", "SP", "PR", "SC", "RS", "MS",
    "MT", "GO", "DF",
]

REGIAO_POR_UF = {
    "RO": "Norte", "AC": "Norte", "AM": "Norte", "RR": "Norte", "PA": "Norte",
    "AP": "Norte", "TO": "Norte",
    "MA": "Nordeste", "PI": "Nordeste", "CE": "Nordeste", "RN": "Nordeste",
    "PB": "Nordeste", "PE": "Nordeste", "AL": "Nordeste", "SE": "Nordeste",
    "BA": "Nordeste",
    "MG": "Sudeste", "ES": "Sudeste", "RJ": "Sudeste", "SP": "Sudeste",
    "PR": "Sul", "SC": "Sul", "RS": "Sul",
    "MS": "Centro-Oeste", "MT": "Centro-Oeste", "GO": "Centro-Oeste",
    "DF": "Centro-Oeste",
}

UNIDADES_COMPATIVEIS = {
    "KG", "L", "T", "TON", "TONELADA", "TONELADAS", "LITRO", "LITROS",
    "QUILOGRAMA", "QUILOGRAMAS", "G", "GRAMA", "GRAMAS", "ML", "M3",
}

UNIDADES_CONCENTRACAO = {
    "G_L", "G/L", "GL", "G_KG", "G/KG", "GKG", "MG_L", "MG/L",
    "PERCENT", "PERCENTUAL", "%", "PPM", "PPB",
}

COLUNAS_NUMERICAS_COMERC = [
    "ESTOQUE_INICIAL", "PRODUCAO_NACIONAL", "IMPORTACAO", "EXPORTACAO",
    "VENDAS_INDUSTRIA", "VENDAS_CLIENTE", "VENDAS_TOTAIS", "ESTOQUE_FINAL",
]

MAPA_VOLUMES_PONDERADOS = {
    "ESTOQUE_INICIAL": "ESTOQUE_INICIAL_PONDERADO_IA",
    "PRODUCAO_NACIONAL": "PRODUCAO_NACIONAL_PONDERADA_IA",
    "IMPORTACAO": "IMPORTACAO_PONDERADA_IA",
    "EXPORTACAO": "EXPORTACAO_PONDERADA_IA",
    "VENDAS_INDUSTRIA": "VENDAS_INDUSTRIA_PONDERADA_IA",
    "VENDAS_CLIENTE": "VENDAS_CLIENTE_PONDERADA_IA",
    "VENDAS_TOTAIS": "VENDA_PONDERADA_IA",
    "ESTOQUE_FINAL": "ESTOQUE_FINAL_PONDERADO_IA",
}

LOGGER = logging.getLogger("agrofit_comercializacao")

# =============================================================================
# LOGGING E MENSAGENS
# =============================================================================

class Cores:
    HEADER = "\033[95m"
    AZUL = "\033[94m"
    VERDE = "\033[92m"
    AMARELO = "\033[93m"
    VERMELHO = "\033[91m"
    FIM = "\033[0m"
    NEGRITO = "\033[1m"


if sys.platform == "win32":
    try:
        import ctypes
        kernel32 = ctypes.windll.kernel32
        kernel32.SetConsoleMode(kernel32.GetStdHandle(-11), 7)
    except Exception:
        for _attr in ["HEADER", "AZUL", "VERDE", "AMARELO", "VERMELHO", "FIM", "NEGRITO"]:
            setattr(Cores, _attr, "")


def configurar_logging() -> None:
    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    caminho_log = LOGS_DIR / ARQUIVO_LOG
    LOGGER.setLevel(logging.DEBUG)
    LOGGER.handlers.clear()
    fh = logging.FileHandler(caminho_log, mode="a", encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter(
        "%(asctime)s | %(levelname)-8s | %(message)s", "%Y-%m-%d %H:%M:%S"
    ))
    LOGGER.addHandler(fh)


def titulo(texto: str) -> None:
    linha = "=" * 78
    print(f"\n{Cores.HEADER}{Cores.NEGRITO}{linha}{Cores.FIM}")
    print(f"{Cores.NEGRITO}{texto.center(78)}{Cores.FIM}")
    print(f"{Cores.HEADER}{linha}{Cores.FIM}\n")
    LOGGER.info("=== %s ===", texto)


def ok(msg: str) -> None:
    print(f"{Cores.VERDE}[OK] {msg}{Cores.FIM}")
    LOGGER.info("OK: %s", msg)


def aviso(msg: str) -> None:
    print(f"{Cores.AMARELO}[AVISO] {msg}{Cores.FIM}")
    LOGGER.warning(msg)


def erro(msg: str) -> None:
    print(f"{Cores.VERMELHO}[ERRO] {msg}{Cores.FIM}")
    LOGGER.error(msg)


def passo(msg: str) -> None:
    print(f"{Cores.AZUL}-> {msg}{Cores.FIM}")
    LOGGER.info("PASSO: %s", msg)


def tempo_legivel(segundos: float) -> str:
    if segundos < 60:
        return f"{segundos:.1f}s"
    if segundos < 3600:
        return f"{segundos / 60:.1f} min"
    return f"{segundos / 3600:.2f} h"


class Cronometro:
    def __init__(self, nome: str):
        self.nome = nome
        self.inicio = None

    def __enter__(self):
        self.inicio = time.time()
        passo(f"Iniciando: {self.nome}")
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        duracao = time.time() - self.inicio
        if exc_type is None:
            ok(f"Concluido: {self.nome} ({tempo_legivel(duracao)})")
        else:
            erro(f"Falhou: {self.nome} ({tempo_legivel(duracao)}) -> {exc_val}")
        LOGGER.info("TEMPO | %s | %.2fs", self.nome, duracao)
        return False


# =============================================================================
# NORMALIZACAO
# =============================================================================

def remover_acentos(texto: str) -> str:
    if texto is None:
        return ""
    nfkd = unicodedata.normalize("NFKD", str(texto))
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def normalizar_nome_coluna(nome: str) -> str:
    if nome is None:
        return ""
    s = str(nome)
    s = re.sub(r"<br\s*/?>", " ", s, flags=re.IGNORECASE)
    s = s.replace("\n", " ").replace("\r", " ")
    s = remover_acentos(s).strip().upper()
    s = re.sub(r"[^\w]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def normalizar_texto_valor(valor: Any) -> Any:
    if pd.isna(valor):
        return valor
    s = remover_acentos(str(valor)).strip().upper()
    s = re.sub(r"\s+", " ", s)
    return s


def extrair_chave_registro(valor: Any) -> str:
    if pd.isna(valor):
        return ""
    s = str(valor).strip().upper()
    s = remover_acentos(s)
    s = re.sub(r"^(MAPA|IBAMA|ANVISA|N[R°º]?\.?|REG(?:ISTRO)?|N[ºO]\.?)[\s\-/:]*", "", s)
    s = s.replace(" ", "").replace("_", "").replace("/", "").replace("-", "")
    digitos = re.sub(r"[^0-9A-Z]", "", s)
    only_digits = re.sub(r"\D", "", digitos)
    if not only_digits:
        return ""
    return only_digits.lstrip("0") or "0"


def normalizar_ia(valor: Any) -> Any:
    if pd.isna(valor):
        return valor
    s = remover_acentos(str(valor)).strip().lower()
    s = re.sub(r"\s+", " ", s)
    s_sem_paren = re.sub(r"\([^)]*\)", "", s).strip()
    if s_sem_paren:
        s = s_sem_paren
    s = re.sub(r"\s+", " ", s).strip(" ;,.-")
    return s if s else valor


# =============================================================================
# ALIASES DE COLUNA
# =============================================================================

ALIASES_AGROFIT: Dict[str, List[str]] = {
    "NR_REGISTRO": ["NR_REGISTRO", "NUMERO_REGISTRO", "REGISTRO", "N_REGISTRO",
                     "REGISTRO_MAPA", "NR_REGISTRO_MAPA", "NO_REGISTRO",
                     "NUMERO_DO_REGISTRO", "NO_DE_REGISTRO", "N_DE_REGISTRO"],
    "MARCA_COMERCIAL": ["MARCA_COMERCIAL", "MARCA", "NOME_COMERCIAL", "PRODUTO"],
    "FORMULACAO": ["FORMULACAO", "TIPO_FORMULACAO"],
    "INGREDIENTE_ATIVO": ["INGREDIENTE_ATIVO", "INGREDIENTES_ATIVOS", "IA",
                           "PRINCIPIO_ATIVO"],
    "TITULAR_DE_REGISTRO": ["TITULAR_DE_REGISTRO", "TITULAR_REGISTRO", "TITULAR",
                             "EMPRESA_TITULAR", "DETENTOR_DO_REGISTRO"],
    "CLASSE": ["CLASSE", "CLASSE_DO_PRODUTO", "CLASSE_AGRONOMICA"],
    "MODO_DE_ACAO": ["MODO_DE_ACAO", "MODO_ACAO", "MECANISMO_DE_ACAO"],
    "CULTURA": ["CULTURA", "CULTURAS"],
    "PRAGA_NOME_CIENTIFICO": ["PRAGA_NOME_CIENTIFICO", "NOME_CIENTIFICO_PRAGA",
                               "NOME_CIENTIFICO"],
    "PRAGA_NOME_COMUM": ["PRAGA_NOME_COMUM", "NOME_COMUM_PRAGA", "NOME_COMUM"],
    "EMPRESA_PAIS_TIPO": ["EMPRESA_PAIS_TIPO", "PAIS_TITULAR", "PAIS", "TIPO_EMPRESA"],
    "CLASSE_TOXICOLOGICA": ["CLASSE_TOXICOLOGICA", "CLASSIFICACAO_TOXICOLOGICA",
                             "CLASSE_TOXICOLOGICA_ANVISA"],
    "CLASSE_AMBIENTAL": ["CLASSE_AMBIENTAL", "CLASSIFICACAO_AMBIENTAL",
                          "CLASSE_AMBIENTAL_IBAMA"],
    "ORGANICOS": ["ORGANICOS", "USO_EM_AGRICULTURA_ORGANICA", "AGRICULTURA_ORGANICA"],
    "SITUACAO": ["SITUACAO", "SITUACAO_DO_PRODUTO", "STATUS"],
}

ALIASES_COMERC: Dict[str, List[str]] = {
    "NR_REGISTRO": ["NR_REGISTRO", "NUMERO_REGISTRO", "REGISTRO", "N_REGISTRO",
                     "REGISTRO_MAPA", "REGISTRO_DO_PRODUTO", "N_REGISTRO_MAPA",
                     "NO_REGISTRO", "NUMERO_DO_REGISTRO", "NO_DE_REGISTRO", "N_DE_REGISTRO"],
    "MARCA_COMERCIAL": ["MARCA_COMERCIAL", "PRODUTO_COMERCIAL", "MARCA",
                         "NOME_DO_PRODUTO", "MARCA_COMERCIAL_VENDAS"],
    "INGREDIENTE_ATIVO": ["INGREDIENTE_ATIVO", "INGREDIENTES_ATIVOS", "IA",
                           "PRINCIPIO_ATIVO", "INGREDIENTE_ATIVO_VENDAS"],
    "DETENTOR": ["DETENTOR", "DETENTOR_DO_REGISTRO", "TITULAR_DE_REGISTRO",
                  "TITULAR", "EMPRESA"],
    "PRODUTO_TECNICO": ["PRODUTO_TECNICO", "TIPO_DE_PRODUTO", "CATEGORIA"],
    "CONCENTRACAO": ["CONCENTRACAO", "CONCENTRACAO_DO_IA", "TEOR"],
    "UNIDADE": ["UNIDADE", "UNIDADE_DE_MEDIDA", "UM", "UNID", "UNIDADE_CONCENTRACAO"],
    "CLASSES_USO": ["CLASSES_USO", "CLASSE_DE_USO", "CLASSES_DE_USO"],
    "PONDERACAO_IA": [
        "PONDERACAO_IA", "PONDERACAO", "PONDERACAO_PERCENT_IA_NO_PF",
        "PONDERACAO_DO_IA", "PERCENTUAL_IA", "PONDERACAO_IA_NO_PF",
    ],
    "ESTOQUE_INICIAL": ["ESTOQUE_INICIAL", "ESTOQUE_INICIAL_PF", "ESTOQUE_INI"],
    "PRODUCAO_NACIONAL": ["PRODUCAO_NACIONAL", "PRODUCAO", "PRODUCAO_NACIONAL_PF"],
    "IMPORTACAO": ["IMPORTACAO", "IMPORTACAO_PF", "IMPORTAC"],
    "EXPORTACAO": ["EXPORTACAO", "EXPORTACAO_PF", "EXPORTAC"],
    "VENDAS_INDUSTRIA": [
        "VENDAS_INDUSTRIA", "VENDA_INDUSTRIA", "VENDAS_A_INDUSTRIA",
        "VENDAS_INDUSTRIA_PF", "VENDAS_A_INDUSTRIA_PF", "VENDAS_INDUST",
    ],
    "VENDAS_CLIENTE": [
        "VENDAS_CLIENTE", "VENDA_CLIENTE", "VENDAS_A_CLIENTE",
        "VENDAS_CLIENTE_PF", "VENDAS_A_CLIENTE_PF", "VENDAS_CLIEN",
    ],
    "VENDAS_TOTAIS": [
        "VENDAS_TOTAIS", "VENDA_TOTAL", "TOTAL_DE_VENDAS", "VENDAS",
        "VENDAS_PF_TOTAL", "VENDAS_TOTAIS_PF",
    ],
    "ESTOQUE_FINAL": ["ESTOQUE_FINAL", "ESTOQUE_FINAL_PF", "ESTOQUE_FIN"],
    "ANO": ["ANO", "ANO_BASE", "ANO_REFERENCIA", "ANO_COMERCIALIZACAO"],
}


def padronizar_colunas(
    df: pd.DataFrame, aliases: Dict[str, List[str]], origem: str = ""
) -> pd.DataFrame:
    df = df.copy()
    df.columns = [normalizar_nome_coluna(c) for c in df.columns]
    mapa_reverso: Dict[str, str] = {}
    for padrao, variantes in aliases.items():
        for v in variantes:
            mapa_reverso[v] = padrao
    novas_colunas = {}
    nao_mapeadas = []
    for col in df.columns:
        if col in mapa_reverso:
            novas_colunas[col] = mapa_reverso[col]
        else:
            novas_colunas[col] = col
            if col not in aliases:
                nao_mapeadas.append(col)
    df = df.rename(columns=novas_colunas)
    if df.columns.duplicated().any():
        dup_names = df.columns[df.columns.duplicated()].unique().tolist()
        for nome in dup_names:
            sub = df.loc[:, df.columns == nome]
            combinado = sub.bfill(axis=1).iloc[:, 0]
            df = df.loc[:, df.columns != nome]
            df[nome] = combinado
    if nao_mapeadas:
        LOGGER.info(
            "Colunas nao mapeadas em %s (mantidas): %s",
            origem, ", ".join(sorted(set(nao_mapeadas))),
        )
    return df


# =============================================================================
# CONVERSAO NUMERICA
# =============================================================================

def converter_numero_br(serie: pd.Series) -> pd.Series:
    if serie is None:
        return serie
    if pd.api.types.is_numeric_dtype(serie):
        return pd.to_numeric(serie, errors="coerce")

    def _conv(v):
        if pd.isna(v):
            return np.nan
        if isinstance(v, (int, float, np.integer, np.floating)):
            return float(v)
        s = str(v).strip()
        if s == "" or s.upper() in {"NA", "N/A", "-", "ND", "NAO_INFORMADO"}:
            return np.nan
        s = re.sub(r"[^\d,.\-]", "", s)
        if s == "":
            return np.nan
        if "," in s and "." in s:
            if s.rfind(",") > s.rfind("."):
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        elif "," in s:
            s = s.replace(".", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return np.nan

    return serie.apply(_conv)


def converter_percentual(serie: pd.Series) -> pd.Series:
    valores = converter_numero_br(serie)

    def _norm(v):
        if pd.isna(v):
            return np.nan
        v = float(v)
        if v > 1.0:
            v = v / 100.0
        if v < 0 or v > 1:
            return np.nan
        return v

    return valores.apply(_norm)


# =============================================================================
# CARGA AGROFIT
# =============================================================================

def localizar_arquivo_agrofit() -> Optional[Path]:
    caminho_padrao = DADOS_DIR / ARQUIVO_AGROFIT_PADRAO
    if caminho_padrao.exists():
        return caminho_padrao
    for arq in DADOS_DIR.glob("*.csv"):
        nome_norm = remover_acentos(arq.stem).lower()
        if any(pista in nome_norm for pista in PISTAS_NOME_AGROFIT):
            return arq
    return None


def _ler_csv_robusto(caminho: Path) -> pd.DataFrame:
    tentativas = [
        dict(sep=";", encoding="utf-8-sig"),
        dict(sep=";", encoding="latin1"),
        dict(sep=",", encoding="utf-8-sig"),
        dict(sep=",", encoding="latin1"),
        dict(sep="\t", encoding="utf-8-sig"),
    ]
    ultimo_erro = None
    for opcoes in tentativas:
        try:
            df = pd.read_csv(caminho, dtype=str, low_memory=False, **opcoes)
            if df.shape[1] > 1:
                return df
        except Exception as e:
            ultimo_erro = e
            continue
    raise RuntimeError(f"Nao foi possivel ler o CSV {caminho.name}: {ultimo_erro}")


def carregar_agrofit(caminho: Path) -> pd.DataFrame:
    df = _ler_csv_robusto(caminho)
    linhas_brutas = len(df)
    df = padronizar_colunas(df, ALIASES_AGROFIT, origem=f"Agrofit ({caminho.name})")
    for col in ["NR_REGISTRO", "MARCA_COMERCIAL", "INGREDIENTE_ATIVO", "CULTURA",
                "PRAGA_NOME_CIENTIFICO", "PRAGA_NOME_COMUM", "TITULAR_DE_REGISTRO",
                "FORMULACAO", "SITUACAO"]:
        if col not in df.columns:
            df[col] = np.nan
            aviso(f"Coluna esperada do Agrofit ausente: {col}")
    df["NR_REGISTRO"] = df["NR_REGISTRO"].astype(str).str.strip()
    df["_CHAVE_REGISTRO"] = df["NR_REGISTRO"].apply(extrair_chave_registro)
    df = df[df["_CHAVE_REGISTRO"] != ""]
    LOGGER.info(
        "Agrofit: %s linhas brutas, %s validas, %s registros unicos",
        linhas_brutas, len(df), df["_CHAVE_REGISTRO"].nunique(),
    )
    return df


# =============================================================================
# CARGA COMERCIALIZACAO
# =============================================================================

PADRAO_ANO = re.compile(r"(?:19|20)\d{2}")


def extrair_ano_do_nome(nome_arquivo: str) -> Optional[int]:
    achados = re.findall(r"(?:19|20)\d{2}", nome_arquivo)
    if not achados:
        return None
    return int(achados[-1])


def detectar_arquivos_comercializacao(caminho_agrofit: Optional[Path]) -> List[Path]:
    candidatos: List[Path] = []
    for ext in ["*.xlsx", "*.xls", "*.csv"]:
        candidatos.extend(DADOS_DIR.glob(ext))
    candidatos = [
        c for c in candidatos
        if not (caminho_agrofit is not None and c.resolve() == caminho_agrofit.resolve())
        and not c.name.startswith("~$")
    ]
    return sorted(candidatos)


def _escolher_aba_excel(caminho: Path) -> Optional[str]:
    try:
        xls = pd.ExcelFile(caminho)
    except Exception:
        return None
    abas = xls.sheet_names
    if len(abas) == 1:
        return abas[0]
    prioridade = ["TODOS_AGROTOXICOS", "TODOS", "BASE", "DADOS", "TABELA_BRUTA", "COMERCIALIZACAO"]
    for pista in prioridade:
        for aba in abas:
            if pista in remover_acentos(aba).upper().replace(" ", "_"):
                return aba
    melhor_aba, melhor_cols = abas[0], -1
    for aba in abas:
        try:
            df_teste = pd.read_excel(caminho, sheet_name=aba, nrows=5)
            if df_teste.shape[1] > melhor_cols:
                melhor_cols = df_teste.shape[1]
                melhor_aba = aba
        except Exception:
            continue
    return melhor_aba


def _limpar_cabecalho_tabela_bruta(nome: str) -> str:
    if nome is None:
        return ""
    s = str(nome).replace("\xa0", " ").strip()
    s = re.sub(r"<br\s*/?>", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _mapear_colunas_tabela_bruta(df: pd.DataFrame) -> pd.DataFrame:
    """Mapeia a estrutura da Tabela Bruta IBAMA/MAPA (bloco PF + bloco IA)."""
    df = df.copy()
    df.columns = [_limpar_cabecalho_tabela_bruta(c) for c in df.columns]
    df = df.loc[:, [c for c in df.columns if c and not str(c).startswith("Unnamed")]]

    rename: Dict[str, str] = {}
    cols = list(df.columns)

    meta = {
        "Detentor": "DETENTOR",
        "Marca comercial": "MARCA_COMERCIAL",
        "Tipo": "PRODUTO_TECNICO",
        "Registro Mapa": "NR_REGISTRO",
        "Classes Uso": "CLASSES_USO",
        "Produto Tecnico": "PRODUTO_TECNICO_DETALHE",
        "Produto Técnico": "PRODUTO_TECNICO_DETALHE",
        "Ingrediente Ativo": "INGREDIENTE_ATIVO",
        "Concentracao": "CONCENTRACAO",
        "Concentração": "CONCENTRACAO",
        "Unidade": "UNIDADE",
        "Ponderacao (% IA no PF)": "PONDERACAO_IA",
        "Ponderação (% IA no PF)": "PONDERACAO_IA",
        "Linha": "LINHA",
        "Densidade": "DENSIDADE",
        "Total UF": "TOTAL_UF_IA",
        "Brasil": "BRASIL_IA",
        "Vendas Totais": "VENDAS_TOTAIS_IA_ORIGEM",
    }
    for c in cols:
        if c in meta:
            rename[c] = meta[c]
        c_sem = remover_acentos(c)
        if c_sem in meta and c not in rename:
            rename[c] = meta[c_sem]

    pf_map = {
        "Estoque Inicial": "ESTOQUE_INICIAL",
        "Producao Nacional": "PRODUCAO_NACIONAL",
        "Produção Nacional": "PRODUCAO_NACIONAL",
        "Importacao": "IMPORTACAO",
        "Importação": "IMPORTACAO",
        "Exportacao": "EXPORTACAO",
        "Exportação": "EXPORTACAO",
        "Vendas a Industria": "VENDAS_INDUSTRIA",
        "Vendas a Indústria": "VENDAS_INDUSTRIA",
        "Vendas a Cliente": "VENDAS_CLIENTE",
        "Estoque Final": "ESTOQUE_FINAL",
    }
    ia_map = {
        "Estoque Inicial.1": "ESTOQUE_INICIAL_PONDERADO_IA",
        "Producao Nacional.1": "PRODUCAO_NACIONAL_PONDERADA_IA",
        "Produção Nacional.1": "PRODUCAO_NACIONAL_PONDERADA_IA",
        "Importacao.1": "IMPORTACAO_PONDERADA_IA",
        "Importação.1": "IMPORTACAO_PONDERADA_IA",
        "Exportacao.1": "EXPORTACAO_PONDERADA_IA",
        "Exportação.1": "EXPORTACAO_PONDERADA_IA",
        "Vendas a Industria.1": "VENDAS_INDUSTRIA_PONDERADA_IA",
        "Vendas a Indústria.1": "VENDAS_INDUSTRIA_PONDERADA_IA",
        "Vendas a Cliente.1": "VENDAS_CLIENTE_PONDERADA_IA",
        "Estoque Final.1": "ESTOQUE_FINAL_PONDERADO_IA",
    }

    for c in cols:
        if c in pf_map:
            rename[c] = pf_map[c]
        elif c in ia_map:
            rename[c] = ia_map[c]
        else:
            if c in UFS:
                rename[c] = c
            elif c.endswith(".1") and c[:-2] in UFS:
                rename[c] = f"UF_{c[:-2]}_IA"

    df = df.rename(columns=rename)
    if df.columns.duplicated().any():
        dup_names = df.columns[df.columns.duplicated()].unique().tolist()
        for nome in dup_names:
            sub = df.loc[:, df.columns == nome]
            combinado = sub.bfill(axis=1).iloc[:, 0]
            df = df.loc[:, df.columns != nome]
            df[nome] = combinado
    return df


def carregar_arquivo_comercializacao(caminho: Path) -> pd.DataFrame:
    ano_detectado = extrair_ano_do_nome(caminho.name)

    if caminho.suffix.lower() == ".csv":
        df = _ler_csv_robusto(caminho)
        df = padronizar_colunas(df, ALIASES_COMERC, origem=f"Comercializacao ({caminho.name})")
    else:
        aba = _escolher_aba_excel(caminho)
        df = pd.read_excel(caminho, sheet_name=aba, dtype=str)
        cols_raw = [str(c) for c in df.columns]
        is_tabela_bruta = any(
            ("<br>" in c) or ("Pondera" in c) or (c.strip() == "Brasil")
            for c in cols_raw
        )
        if is_tabela_bruta:
            LOGGER.info("Tabela Bruta detectada em %s (aba=%s)", caminho.name, aba)
            df = _mapear_colunas_tabela_bruta(df)
            df = padronizar_colunas(df, ALIASES_COMERC, origem=f"Comercializacao/TB ({caminho.name})")
        else:
            df = padronizar_colunas(df, ALIASES_COMERC, origem=f"Comercializacao ({caminho.name})")

    if "ANO" in df.columns and df["ANO"].notna().any():
        df["ANO"] = converter_numero_br(df["ANO"]).astype("Int64")
        if ano_detectado is not None and df["ANO"].isna().all():
            df["ANO"] = ano_detectado
    elif ano_detectado is not None:
        df["ANO"] = ano_detectado
    else:
        aviso(f"ANO nao identificado em {caminho.name}; linhas sem serie historica.")
        df["ANO"] = np.nan

    df["_ARQUIVO_ORIGEM"] = caminho.name

    if "NR_REGISTRO" not in df.columns:
        candidatas = [
            c for c in df.columns
            if "REGISTRO" in c and not any(p in c for p in ["SITUACAO", "TITULAR", "DETENTOR", "DATA"])
        ]
        if candidatas:
            df = df.rename(columns={candidatas[0]: "NR_REGISTRO"})
            LOGGER.info("Coluna '%s' assumida como NR_REGISTRO em %s.", candidatas[0], caminho.name)
        else:
            aviso(f"{caminho.name} sem NR_REGISTRO identificavel; ignorado.")
            return pd.DataFrame()

    df["NR_REGISTRO"] = df["NR_REGISTRO"].astype(str).str.strip()
    df["_CHAVE_REGISTRO"] = df["NR_REGISTRO"].apply(extrair_chave_registro)
    df = df[df["_CHAVE_REGISTRO"] != ""]

    LOGGER.info(
        "Comercializacao: %s | linhas=%s | anos=%s",
        caminho.name, len(df), sorted(df["ANO"].dropna().unique().tolist()),
    )
    return df


def carregar_todas_comercializacoes(arquivos: List[Path]) -> pd.DataFrame:
    partes = []
    for caminho in arquivos:
        try:
            df = carregar_arquivo_comercializacao(caminho)
            if not df.empty:
                partes.append(df)
                ok(f"Processado: {caminho.name} ({len(df):,} linhas)")
            else:
                aviso(f"Arquivo ignorado (sem dados): {caminho.name}")
        except Exception as e:
            erro(f"Falha ao processar {caminho.name}: {e}")
            LOGGER.error("Falha %s: %s\n%s", caminho.name, e, traceback.format_exc())
    if not partes:
        return pd.DataFrame()
    return pd.concat(partes, ignore_index=True, sort=False)


# =============================================================================
# PREPARACAO NUMERICA E PONDERACAO IA
# =============================================================================

def preparar_numericos_comercializacao(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in COLUNAS_NUMERICAS_COMERC:
        if col in df.columns:
            df[col] = converter_numero_br(df[col])
        else:
            df[col] = np.nan

    extras_num = [
        "PONDERACAO_IA", "VENDAS_TOTAIS_IA_ORIGEM", "BRASIL_IA", "TOTAL_UF_IA",
        "CONCENTRACAO", "DENSIDADE",
    ] + list(MAPA_VOLUMES_PONDERADOS.values())
    for col in extras_num:
        if col in df.columns:
            if col == "PONDERACAO_IA":
                df[col] = converter_percentual(df[col])
            else:
                df[col] = converter_numero_br(df[col])

    if "PONDERACAO_IA" not in df.columns:
        df["PONDERACAO_IA"] = np.nan

    soma_ind_cli = df[["VENDAS_INDUSTRIA", "VENDAS_CLIENTE"]].sum(axis=1, min_count=1)
    if "VENDAS_TOTAIS" not in df.columns or df["VENDAS_TOTAIS"].isna().all():
        df["VENDAS_TOTAIS"] = soma_ind_cli
    else:
        df["VENDAS_TOTAIS"] = df["VENDAS_TOTAIS"].where(df["VENDAS_TOTAIS"].notna(), soma_ind_cli)
        if "VENDAS_TOTAIS_IA_ORIGEM" in df.columns:
            df["VENDAS_TOTAIS"] = soma_ind_cli

    if "UNIDADE" not in df.columns:
        df["UNIDADE"] = np.nan
    df["_UNIDADE_NORM"] = df["UNIDADE"].apply(
        lambda v: normalizar_texto_valor(v) if pd.notna(v) else np.nan
    )
    return df


def calcular_venda_estimada_ia(df: pd.DataFrame) -> pd.DataFrame:
    """Preserva volumes ORIGINAIS e cria colunas PONDERADAS (PF x teor IA)."""
    df = df.copy()

    def _status(row):
        vendas = row.get("VENDAS_TOTAIS", np.nan)
        if pd.isna(vendas) and all(
            pd.isna(row.get(c, np.nan)) for c in COLUNAS_NUMERICAS_COMERC if c != "VENDAS_TOTAIS"
        ):
            return "DADO_AUSENTE"
        if pd.isna(row.get("PONDERACAO_IA", np.nan)):
            return "SEM_PONDERACAO"
        unidade = row.get("_UNIDADE_NORM")
        if pd.notna(unidade):
            u = str(unidade).replace("/", "_").replace(" ", "_")
            if u in UNIDADES_CONCENTRACAO or any(
                u.startswith(p) or u.endswith(p) for p in ("G_L", "G_KG", "MG_L")
            ):
                return "CALCULADO"
            if u not in UNIDADES_COMPATIVEIS and u not in {"", "NAN"}:
                if u in {"UN", "UND", "UNIDADE", "PC", "PÇ", "CX", "FR", "SC", "DOSE"}:
                    return "UNIDADE_INCOMPATIVEL"
        return "CALCULADO"

    df["STATUS_CALCULO_IA"] = df.apply(_status, axis=1)
    pode_calcular = df["STATUS_CALCULO_IA"] == "CALCULADO"

    for col_orig, col_pond in MAPA_VOLUMES_PONDERADOS.items():
        if col_orig not in df.columns:
            df[col_orig] = np.nan
        if col_pond in df.columns and df[col_pond].notna().any():
            calculado = np.where(
                pode_calcular & df[col_orig].notna(),
                df[col_orig] * df["PONDERACAO_IA"],
                np.nan,
            )
            df[col_pond] = df[col_pond].where(df[col_pond].notna(), calculado)
        else:
            df[col_pond] = np.where(
                pode_calcular & df.get(col_orig, pd.Series(np.nan, index=df.index)).notna(),
                df[col_orig] * df["PONDERACAO_IA"],
                np.nan,
            )

    if "VENDAS_TOTAIS_IA_ORIGEM" in df.columns and df["VENDAS_TOTAIS_IA_ORIGEM"].notna().any():
        df["VENDA_PONDERADA_IA"] = df["VENDAS_TOTAIS_IA_ORIGEM"].where(
            df["VENDAS_TOTAIS_IA_ORIGEM"].notna(),
            df.get("VENDA_PONDERADA_IA"),
        )
    elif "BRASIL_IA" in df.columns and df["BRASIL_IA"].notna().any():
        df["VENDA_PONDERADA_IA"] = df["BRASIL_IA"].where(
            df["BRASIL_IA"].notna(),
            df.get("VENDA_PONDERADA_IA"),
        )

    if "VENDA_PONDERADA_IA" in df.columns:
        df["VENDA_ESTIMADA_IA"] = df["VENDA_PONDERADA_IA"]
    else:
        df["VENDA_ESTIMADA_IA"] = np.nan
        df["VENDA_PONDERADA_IA"] = np.nan

    return df


# =============================================================================
# UF LONGO
# =============================================================================

def identificar_colunas_uf(df: pd.DataFrame) -> List[str]:
    return [uf for uf in UFS if uf in df.columns]


def montar_venda_para_uf_longo(
    df_vendas_bruto: pd.DataFrame, colunas_id: List[str]
) -> pd.DataFrame:
    colunas_uf = identificar_colunas_uf(df_vendas_bruto)
    if not colunas_uf:
        return pd.DataFrame(columns=["NR_REGISTRO", "ANO", "UF", "REGIAO", "VENDA"])
    df = df_vendas_bruto.copy()
    for uf in colunas_uf:
        df[uf] = converter_numero_br(df[uf])
    cols_manter = [c for c in colunas_id if c in df.columns] + colunas_uf
    df_long = df[cols_manter].melt(
        id_vars=[c for c in colunas_id if c in df.columns],
        value_vars=colunas_uf,
        var_name="UF",
        value_name="VENDA",
    )
    df_long = df_long[df_long["VENDA"].notna() & (df_long["VENDA"] != 0)]
    df_long["REGIAO"] = df_long["UF"].map(REGIAO_POR_UF)
    return df_long


def _enriquecer_com_agrofit(
    df: pd.DataFrame, df_produtos: pd.DataFrame, campos: Optional[List[str]] = None
) -> pd.DataFrame:
    if df.empty or df_produtos.empty:
        return df
    if campos is None:
        campos = ["MARCA_COMERCIAL", "INGREDIENTE_ATIVO", "TITULAR_DE_REGISTRO", "FORMULACAO"]
    campos = [c for c in campos if c in df_produtos.columns]
    if not campos or "_CHAVE_REGISTRO" not in df.columns:
        return df
    df_prod_aux = df_produtos[["_CHAVE_REGISTRO"] + campos].drop_duplicates("_CHAVE_REGISTRO")
    df_prod_aux = df_prod_aux.rename(columns={c: f"{c}_AGROFIT" for c in campos})
    out = df.merge(df_prod_aux, on="_CHAVE_REGISTRO", how="left")
    for c in campos:
        col_agro = f"{c}_AGROFIT"
        if c in out.columns:
            out[c] = out[c].where(
                out[c].notna() & (out[c].astype(str).str.strip() != ""),
                out[col_agro],
            )
        else:
            out[c] = out[col_agro]
        out.drop(columns=[col_agro], inplace=True, errors="ignore")
    return out


def _juntar_unicos(serie: pd.Series) -> str:
    vals = [str(v).strip() for v in serie.dropna().unique() if str(v).strip()]
    return "; ".join(sorted(set(vals))) if vals else np.nan


# =============================================================================
# 01 - RESUMO AGROFIT (dimensao produto)
# =============================================================================

def montar_resumo_agrofit(df_agrofit: pd.DataFrame) -> pd.DataFrame:
    """1 linha por registro MAPA: cadastro Agrofit + contagens de uso."""
    if df_agrofit.empty:
        return pd.DataFrame()

    colunas_attr = [
        "NR_REGISTRO", "MARCA_COMERCIAL", "FORMULACAO", "INGREDIENTE_ATIVO",
        "TITULAR_DE_REGISTRO", "CLASSE", "MODO_DE_ACAO", "EMPRESA_PAIS_TIPO",
        "CLASSE_TOXICOLOGICA", "CLASSE_AMBIENTAL", "ORGANICOS", "SITUACAO",
    ]
    presentes = [c for c in colunas_attr if c in df_agrofit.columns]

    agregadores: Dict[str, Any] = {}
    for c in presentes:
        if c == "NR_REGISTRO":
            continue
        if c == "INGREDIENTE_ATIVO":
            agregadores[c] = _juntar_unicos
        else:
            agregadores[c] = "first"

    df = df_agrofit.copy()
    df_prod = df.groupby(["_CHAVE_REGISTRO", "NR_REGISTRO"], as_index=False).agg(agregadores)
    df_prod = df_prod.drop_duplicates(subset=["_CHAVE_REGISTRO"]).reset_index(drop=True)

    # Contagens de uso
    if "CULTURA" in df.columns:
        uso_c = df.groupby("_CHAVE_REGISTRO")["CULTURA"].nunique().rename("NUM_CULTURAS")
        df_prod = df_prod.merge(uso_c, on="_CHAVE_REGISTRO", how="left")
    else:
        df_prod["NUM_CULTURAS"] = 0

    if "PRAGA_NOME_CIENTIFICO" in df.columns:
        uso_p = df.groupby("_CHAVE_REGISTRO")["PRAGA_NOME_CIENTIFICO"].nunique().rename("NUM_PRAGAS")
        df_prod = df_prod.merge(uso_p, on="_CHAVE_REGISTRO", how="left")
    else:
        df_prod["NUM_PRAGAS"] = 0

    df_prod["NUM_CULTURAS"] = df_prod["NUM_CULTURAS"].fillna(0).astype(int)
    df_prod["NUM_PRAGAS"] = df_prod["NUM_PRAGAS"].fillna(0).astype(int)

    # Principais culturas / pragas (resumo textual limitado)
    if "CULTURA" in df.columns:
        def _top_culturas(s):
            vals = [str(v).strip() for v in s.dropna() if str(v).strip()]
            unicos = sorted(set(vals))[:15]
            return "; ".join(unicos) if unicos else np.nan
        cult = df.groupby("_CHAVE_REGISTRO")["CULTURA"].agg(_top_culturas).rename("PRINCIPAIS_CULTURAS")
        df_prod = df_prod.merge(cult, on="_CHAVE_REGISTRO", how="left")

    if "PRAGA_NOME_COMUM" in df.columns:
        def _top_pragas(s):
            vals = [str(v).strip() for v in s.dropna() if str(v).strip()]
            unicos = sorted(set(vals))[:15]
            return "; ".join(unicos) if unicos else np.nan
        prag = df.groupby("_CHAVE_REGISTRO")["PRAGA_NOME_COMUM"].agg(_top_pragas).rename("PRINCIPAIS_PRAGAS")
        df_prod = df_prod.merge(prag, on="_CHAVE_REGISTRO", how="left")

    return df_prod


# =============================================================================
# 02 - COMERCIALIZACAO_PRODUTOS (Registro x Ano, volumes PF)
# =============================================================================

def montar_comercializacao_produtos(
    df_vendas: pd.DataFrame, df_resumo_agrofit: pd.DataFrame
) -> pd.DataFrame:
    """1 linha = Registro x Ano. Volumes PF originais. Multi-IA: first/max (nao soma)."""
    if df_vendas.empty:
        return pd.DataFrame()

    df = df_vendas.copy()
    for col in COLUNAS_NUMERICAS_COMERC:
        if col in df.columns and not pd.api.types.is_numeric_dtype(df[col]):
            df[col] = converter_numero_br(df[col])
    if "VENDAS_TOTAIS" not in df.columns:
        df["VENDAS_TOTAIS"] = np.nan
    if df["VENDAS_TOTAIS"].isna().all() and {"VENDAS_INDUSTRIA", "VENDAS_CLIENTE"}.issubset(df.columns):
        df["VENDAS_TOTAIS"] = df[["VENDAS_INDUSTRIA", "VENDAS_CLIENTE"]].sum(axis=1, min_count=1)

    campos_volume_pf = [c for c in COLUNAS_NUMERICAS_COMERC if c in df.columns]
    campos_primeiro = [
        c for c in [
            "MARCA_COMERCIAL", "DETENTOR", "PRODUTO_TECNICO",
            "UNIDADE", "CLASSES_USO", "_UNIDADE_NORM", "_ARQUIVO_ORIGEM",
        ] if c in df.columns
    ]

    # first nos volumes PF: multi-IA repete o mesmo volume — nao somar
    agregadores: Dict[str, Any] = {c: "first" for c in campos_volume_pf}
    agregadores.update({c: "first" for c in campos_primeiro})
    if "INGREDIENTE_ATIVO" in df.columns:
        agregadores["INGREDIENTE_ATIVO"] = _juntar_unicos
    if "CONCENTRACAO" in df.columns:
        agregadores["CONCENTRACAO"] = _juntar_unicos
    if "PONDERACAO_IA" in df.columns:
        agregadores["PONDERACAO_IA"] = _juntar_unicos
    if "NR_REGISTRO" in df.columns:
        agregadores["NR_REGISTRO"] = "first"
    agregadores["NUM_LINHAS_ORIGEM"] = ("_CHAVE_REGISTRO", "size")

    group_keys = ["_CHAVE_REGISTRO", "ANO"]
    df_anual = df.groupby(group_keys, as_index=False).agg(
        **{k: v if isinstance(v, tuple) else (k, v) for k, v in {
            **{c: (c, "first") for c in campos_volume_pf},
            **{c: (c, "first") for c in campos_primeiro},
            **({"INGREDIENTE_ATIVO": ("INGREDIENTE_ATIVO", _juntar_unicos)} if "INGREDIENTE_ATIVO" in df.columns else {}),
            **({"CONCENTRACAO": ("CONCENTRACAO", _juntar_unicos)} if "CONCENTRACAO" in df.columns else {}),
            **({"NR_REGISTRO": ("NR_REGISTRO", "first")} if "NR_REGISTRO" in df.columns else {}),
            "NUM_LINHAS_ORIGEM": ("_CHAVE_REGISTRO", "size"),
        }.items()}
    )

    df_anual = _enriquecer_com_agrofit(df_anual, df_resumo_agrofit)
    if "DETENTOR" not in df_anual.columns:
        df_anual["DETENTOR"] = np.nan
    if "INGREDIENTE_ATIVO" in df_anual.columns:
        df_anual["NUM_IAS"] = df_anual["INGREDIENTE_ATIVO"].apply(
            lambda v: len([x for x in str(v).split(";") if x.strip()]) if pd.notna(v) else 0
        )
    else:
        df_anual["NUM_IAS"] = 0

    colunas_finais = [
        "NR_REGISTRO", "ANO", "MARCA_COMERCIAL", "INGREDIENTE_ATIVO", "NUM_IAS",
        "CONCENTRACAO", "DETENTOR", "PRODUTO_TECNICO", "UNIDADE", "CLASSES_USO",
        "ESTOQUE_INICIAL", "PRODUCAO_NACIONAL", "IMPORTACAO", "EXPORTACAO",
        "VENDAS_INDUSTRIA", "VENDAS_CLIENTE", "VENDAS_TOTAIS", "ESTOQUE_FINAL",
        "NUM_LINHAS_ORIGEM", "_CHAVE_REGISTRO",
    ]
    colunas_finais = [c for c in colunas_finais if c in df_anual.columns]
    return df_anual[colunas_finais].sort_values(["NR_REGISTRO", "ANO"]).reset_index(drop=True)


# =============================================================================
# 03 - EVOLUCAO_PRODUTOS (wide produto x anos)
# =============================================================================

def montar_evolucao_produtos(df_com_produtos: pd.DataFrame) -> pd.DataFrame:
    """1 linha por produto; colunas = anos + TOTAL_PERIODO (volume PF)."""
    if (
        df_com_produtos.empty
        or "ANO" not in df_com_produtos.columns
        or "VENDAS_TOTAIS" not in df_com_produtos.columns
    ):
        return pd.DataFrame()

    df = df_com_produtos.copy()
    df["ANO"] = pd.to_numeric(df["ANO"], errors="coerce").astype("Int64")
    df["VENDAS_TOTAIS"] = pd.to_numeric(df["VENDAS_TOTAIS"], errors="coerce")
    df = df[df["ANO"].notna()]
    if df.empty:
        return pd.DataFrame()

    campos_id = [c for c in ["NR_REGISTRO", "MARCA_COMERCIAL"] if c in df.columns]
    if not campos_id:
        return pd.DataFrame()

    agg = df.groupby(campos_id + ["ANO"], as_index=False)["VENDAS_TOTAIS"].sum()
    pivot = agg.pivot_table(
        index=campos_id, columns="ANO", values="VENDAS_TOTAIS", aggfunc="sum", fill_value=0
    )
    anos_ordenados = sorted(pivot.columns.tolist())
    pivot = pivot[anos_ordenados]
    colunas_ano = [str(int(a)) for a in anos_ordenados]
    pivot.columns = colunas_ano
    pivot["TOTAL_PERIODO"] = pivot[colunas_ano].sum(axis=1)
    pivot = pivot.reset_index()
    return pivot.sort_values("TOTAL_PERIODO", ascending=False).reset_index(drop=True)


# =============================================================================
# 04 - COMERCIALIZACAO_IA (Registro x Ano x IA)
# =============================================================================

def montar_comercializacao_ia(
    df_vendas: pd.DataFrame, df_resumo_agrofit: pd.DataFrame
) -> pd.DataFrame:
    """1 linha = Registro x Ano x IA. Originais PF + ponderados desta IA."""
    if df_vendas.empty:
        return pd.DataFrame()

    df = calcular_venda_estimada_ia(df_vendas.copy())
    df = _enriquecer_com_agrofit(
        df, df_resumo_agrofit, campos=["MARCA_COMERCIAL", "TITULAR_DE_REGISTRO", "FORMULACAO"]
    )
    if "DETENTOR" not in df.columns:
        df["DETENTOR"] = np.nan
    if "VENDA_PONDERADA_IA" in df.columns:
        df["VENDA_ESTIMADA_IA"] = df["VENDA_PONDERADA_IA"]
    elif "VENDA_ESTIMADA_IA" in df.columns:
        df["VENDA_PONDERADA_IA"] = df["VENDA_ESTIMADA_IA"]

    colunas_finais = [
        "NR_REGISTRO", "ANO", "MARCA_COMERCIAL", "INGREDIENTE_ATIVO",
        "DETENTOR", "PRODUTO_TECNICO", "CONCENTRACAO", "UNIDADE", "CLASSES_USO",
        "PONDERACAO_IA",
        # Originais PF da linha (iguais entre IAs do mesmo produto — NAO somar entre IAs)
        "ESTOQUE_INICIAL", "PRODUCAO_NACIONAL", "IMPORTACAO", "EXPORTACAO",
        "VENDAS_INDUSTRIA", "VENDAS_CLIENTE", "VENDAS_TOTAIS", "ESTOQUE_FINAL",
        # Ponderados DESTA IA
        "ESTOQUE_INICIAL_PONDERADO_IA", "PRODUCAO_NACIONAL_PONDERADA_IA",
        "IMPORTACAO_PONDERADA_IA", "EXPORTACAO_PONDERADA_IA",
        "VENDAS_INDUSTRIA_PONDERADA_IA", "VENDAS_CLIENTE_PONDERADA_IA",
        "VENDA_PONDERADA_IA", "VENDA_ESTIMADA_IA", "ESTOQUE_FINAL_PONDERADO_IA",
        "STATUS_CALCULO_IA", "_CHAVE_REGISTRO",
    ]
    colunas_finais = [c for c in colunas_finais if c in df.columns]
    return (
        df[colunas_finais]
        .sort_values(["NR_REGISTRO", "ANO", "INGREDIENTE_ATIVO"])
        .reset_index(drop=True)
    )


# =============================================================================
# 05 - VENDAS_IA (IA x Ano, agregacao ponderada)
# =============================================================================

def montar_vendas_ia(df_com_ia: pd.DataFrame) -> pd.DataFrame:
    """1 linha = IA x Ano. Soma dos volumes PONDERADOS. Contagens de produtos/registros."""
    if df_com_ia.empty or "INGREDIENTE_ATIVO" not in df_com_ia.columns:
        return pd.DataFrame()

    df = df_com_ia.copy()
    df = df[df["INGREDIENTE_ATIVO"].notna() & (df["INGREDIENTE_ATIVO"].astype(str).str.strip() != "")]
    df["_IA_NORM"] = df["INGREDIENTE_ATIVO"].apply(normalizar_ia)
    canon = df.groupby("_IA_NORM")["INGREDIENTE_ATIVO"].agg(lambda s: s.value_counts().index[0])
    df["INGREDIENTE_ATIVO"] = df["_IA_NORM"].map(canon)
    df = df.drop(columns=["_IA_NORM"])

    agg_kwargs: Dict[str, Any] = {
        "NUM_REGISTROS": ("NR_REGISTRO", "nunique"),
        "VENDA_TOTAL_PF_ASSOCIADA": ("VENDAS_TOTAIS", "sum"),
        "VENDA_PONDERADA_IA": ("VENDA_PONDERADA_IA", "sum"),
        "VENDA_ESTIMADA_IA": ("VENDA_ESTIMADA_IA", "sum"),
        "PRODUCAO_NACIONAL_PONDERADA_IA": ("PRODUCAO_NACIONAL_PONDERADA_IA", "sum"),
        "IMPORTACAO_PONDERADA_IA": ("IMPORTACAO_PONDERADA_IA", "sum"),
        "EXPORTACAO_PONDERADA_IA": ("EXPORTACAO_PONDERADA_IA", "sum"),
        "VENDAS_INDUSTRIA_PONDERADA_IA": ("VENDAS_INDUSTRIA_PONDERADA_IA", "sum"),
        "VENDAS_CLIENTE_PONDERADA_IA": ("VENDAS_CLIENTE_PONDERADA_IA", "sum"),
        "ESTOQUE_INICIAL_PONDERADO_IA": ("ESTOQUE_INICIAL_PONDERADO_IA", "sum"),
        "ESTOQUE_FINAL_PONDERADO_IA": ("ESTOQUE_FINAL_PONDERADO_IA", "sum"),
    }
    if "MARCA_COMERCIAL" in df.columns:
        agg_kwargs["NUM_PRODUTOS"] = ("MARCA_COMERCIAL", "nunique")
    else:
        agg_kwargs["NUM_PRODUTOS"] = ("NR_REGISTRO", "nunique")

    agg_kwargs = {
        k: v for k, v in agg_kwargs.items()
        if v[0] in df.columns or k in ("NUM_REGISTROS", "NUM_PRODUTOS")
    }
    grp = df.groupby(["ANO", "INGREDIENTE_ATIVO"], as_index=False).agg(**agg_kwargs)

    if "VENDA_PONDERADA_IA" in grp.columns:
        grp["VENDA_TOTAL"] = grp["VENDA_PONDERADA_IA"]
    elif "VENDA_TOTAL_PF_ASSOCIADA" in grp.columns:
        grp["VENDA_TOTAL"] = grp["VENDA_TOTAL_PF_ASSOCIADA"]

    return grp.sort_values(["ANO", "VENDA_TOTAL"], ascending=[True, False]).reset_index(drop=True)


# =============================================================================
# 06 - EVOLUCAO_IA (wide IA x anos)
# =============================================================================

def montar_evolucao_ia(df_vendas_ia: pd.DataFrame) -> pd.DataFrame:
    """1 linha por IA; colunas = anos + TOTAL_PERIODO (volume ponderado)."""
    if (
        df_vendas_ia.empty
        or "ANO" not in df_vendas_ia.columns
        or "INGREDIENTE_ATIVO" not in df_vendas_ia.columns
    ):
        return pd.DataFrame()

    valor_col = next(
        (
            c for c in ["VENDA_PONDERADA_IA", "VENDA_ESTIMADA_IA", "VENDA_TOTAL", "VENDA_TOTAL_PF_ASSOCIADA"]
            if c in df_vendas_ia.columns
        ),
        None,
    )
    if valor_col is None:
        return pd.DataFrame()

    df = df_vendas_ia.copy()
    df["ANO"] = pd.to_numeric(df["ANO"], errors="coerce").astype("Int64")
    df[valor_col] = pd.to_numeric(df[valor_col], errors="coerce")
    df = df[df["ANO"].notna() & df["INGREDIENTE_ATIVO"].notna()]
    df = df[df["INGREDIENTE_ATIVO"].astype(str).str.strip() != ""]
    if df.empty:
        return pd.DataFrame()

    pivot = df.pivot_table(
        index="INGREDIENTE_ATIVO", columns="ANO", values=valor_col, aggfunc="sum", fill_value=0
    )
    anos_ordenados = sorted(pivot.columns.tolist())
    pivot = pivot[anos_ordenados]
    colunas_ano = [str(int(a)) for a in anos_ordenados]
    pivot.columns = colunas_ano
    pivot["TOTAL_PERIODO"] = pivot[colunas_ano].sum(axis=1)
    pivot = pivot.reset_index()
    return pivot.sort_values("TOTAL_PERIODO", ascending=False).reset_index(drop=True)


# =============================================================================
# 07 - COMERCIALIZACAO_REGIONAL_PRODUTOS (Registro x Ano x UF)
# =============================================================================

def montar_comercializacao_regional_produtos(
    df_vendas: pd.DataFrame, df_resumo_agrofit: Optional[pd.DataFrame] = None
) -> pd.DataFrame:
    """1 linha = Registro x Ano x UF. Volume PF. Multi-IA: max (nao soma)."""
    if df_vendas.empty:
        return pd.DataFrame(columns=[
            "ANO", "NR_REGISTRO", "MARCA_COMERCIAL", "INGREDIENTE_ATIVO",
            "UF", "REGIAO", "VENDA",
        ])

    df_long = montar_venda_para_uf_longo(df_vendas, ["_CHAVE_REGISTRO", "NR_REGISTRO", "ANO"])
    if df_long.empty:
        return df_long

    # max: multi-IA repete o mesmo volume PF
    df_long = df_long.groupby(
        ["_CHAVE_REGISTRO", "NR_REGISTRO", "ANO", "UF", "REGIAO"], as_index=False
    )["VENDA"].max()

    if df_resumo_agrofit is not None and not df_resumo_agrofit.empty:
        campos_id = [c for c in ["MARCA_COMERCIAL", "INGREDIENTE_ATIVO"] if c in df_resumo_agrofit.columns]
        if campos_id:
            df_prod_aux = df_resumo_agrofit[["_CHAVE_REGISTRO"] + campos_id].drop_duplicates("_CHAVE_REGISTRO")
            df_long = df_long.merge(df_prod_aux, on="_CHAVE_REGISTRO", how="left")

    ordem = [
        "ANO", "NR_REGISTRO", "MARCA_COMERCIAL", "INGREDIENTE_ATIVO",
        "UF", "REGIAO", "VENDA", "_CHAVE_REGISTRO",
    ]
    ordem = [c for c in ordem if c in df_long.columns]
    return df_long[ordem].sort_values(["ANO", "NR_REGISTRO", "UF"]).reset_index(drop=True)


# =============================================================================
# 08 - COMERCIALIZACAO_REGIONAL_IA (IA x Ano x UF)
# =============================================================================

def montar_comercializacao_regional_ia(
    df_com_uf: pd.DataFrame, df_com_ia: pd.DataFrame
) -> pd.DataFrame:
    """VENDA_IA_UF = VENDA_PF_UF x PONDERACAO_IA. Agregado por IA x Ano x UF."""
    if df_com_uf.empty or df_com_ia.empty or "INGREDIENTE_ATIVO" not in df_com_ia.columns:
        return pd.DataFrame()

    cols_ia = ["_CHAVE_REGISTRO", "ANO", "INGREDIENTE_ATIVO"]
    if "PONDERACAO_IA" in df_com_ia.columns:
        cols_ia.append("PONDERACAO_IA")
    df_ia = df_com_ia[cols_ia].drop_duplicates()
    df_ia = df_ia[df_ia["INGREDIENTE_ATIVO"].notna()]

    essenciais = ["_CHAVE_REGISTRO", "ANO", "UF", "REGIAO", "VENDA"]
    cols_uf = [c for c in essenciais if c in df_com_uf.columns]
    df_uf = df_com_uf[cols_uf].copy()

    df = df_uf.merge(df_ia, on=["_CHAVE_REGISTRO", "ANO"], how="inner")
    if "PONDERACAO_IA" in df.columns:
        df["VENDA"] = df["VENDA"] * df["PONDERACAO_IA"]

    if "INGREDIENTE_ATIVO" not in df.columns or "VENDA" not in df.columns:
        return pd.DataFrame()

    # Normaliza IA
    df["_IA_NORM"] = df["INGREDIENTE_ATIVO"].apply(normalizar_ia)
    canon = df.groupby("_IA_NORM")["INGREDIENTE_ATIVO"].agg(lambda s: s.value_counts().index[0])
    df["INGREDIENTE_ATIVO"] = df["_IA_NORM"].map(canon)

    grp = df.groupby(["ANO", "INGREDIENTE_ATIVO", "UF", "REGIAO"], as_index=False)["VENDA"].sum()
    return grp.sort_values(
        ["ANO", "INGREDIENTE_ATIVO", "VENDA"], ascending=[True, True, False]
    ).reset_index(drop=True)


# =============================================================================
# COBERTURA / NAO-MATCH
# =============================================================================

def montar_agrofit_sem_vendas(
    df_resumo_agrofit: pd.DataFrame, df_com_produtos: pd.DataFrame
) -> pd.DataFrame:
    chaves_com = set(df_com_produtos["_CHAVE_REGISTRO"].unique()) if not df_com_produtos.empty else set()
    colunas = [
        c for c in [
            "NR_REGISTRO", "MARCA_COMERCIAL", "INGREDIENTE_ATIVO",
            "TITULAR_DE_REGISTRO", "FORMULACAO", "SITUACAO", "_CHAVE_REGISTRO",
        ] if c in df_resumo_agrofit.columns
    ]
    df = df_resumo_agrofit[colunas].copy()
    df = df[~df["_CHAVE_REGISTRO"].isin(chaves_com)]
    df["TEM_COMERCIALIZACAO"] = "NAO"
    return df.drop(columns=["_CHAVE_REGISTRO"], errors="ignore").reset_index(drop=True)


def montar_vendas_sem_agrofit(
    df_com_produtos: pd.DataFrame, df_resumo_agrofit: pd.DataFrame
) -> pd.DataFrame:
    if df_com_produtos.empty:
        return pd.DataFrame()
    chaves_agro = set(df_resumo_agrofit["_CHAVE_REGISTRO"].unique())
    df = df_com_produtos[~df_com_produtos["_CHAVE_REGISTRO"].isin(chaves_agro)].copy()
    df["ENCONTRADO_AGROFIT"] = "NAO"
    return df.drop(columns=["_CHAVE_REGISTRO"], errors="ignore").reset_index(drop=True)


def montar_validacao(
    df_resumo_agrofit: pd.DataFrame,
    df_com_produtos: pd.DataFrame,
    df_agrofit_sem: pd.DataFrame,
    df_vendas_sem: pd.DataFrame,
) -> pd.DataFrame:
    total_agrofit = df_resumo_agrofit["_CHAVE_REGISTRO"].nunique() if "_CHAVE_REGISTRO" in df_resumo_agrofit.columns else 0
    total_com = df_com_produtos["_CHAVE_REGISTRO"].nunique() if not df_com_produtos.empty else 0
    chaves_agro = set(df_resumo_agrofit["_CHAVE_REGISTRO"].unique()) if "_CHAVE_REGISTRO" in df_resumo_agrofit.columns else set()
    chaves_com = set(df_com_produtos["_CHAVE_REGISTRO"].unique()) if not df_com_produtos.empty else set()

    encontrados = len(chaves_agro & chaves_com)
    apenas_agro = len(chaves_agro - chaves_com)
    apenas_com = len(chaves_com - chaves_agro)
    pct = round(100 * encontrados / total_agrofit, 2) if total_agrofit else 0.0

    vol_total = float(df_com_produtos["VENDAS_TOTAIS"].sum()) if not df_com_produtos.empty and "VENDAS_TOTAIS" in df_com_produtos.columns else 0.0
    if not df_com_produtos.empty and "VENDAS_TOTAIS" in df_com_produtos.columns:
        mask_match = df_com_produtos["_CHAVE_REGISTRO"].isin(chaves_agro)
        vol_match = float(df_com_produtos.loc[mask_match, "VENDAS_TOTAIS"].sum())
        vol_sem = float(df_com_produtos.loc[~mask_match, "VENDAS_TOTAIS"].sum())
    else:
        vol_match = 0.0
        vol_sem = 0.0
    pct_vol = round(100 * vol_match / vol_total, 2) if vol_total else 0.0

    linhas = [
        ("TOTAL_REGISTROS_AGROFIT", total_agrofit),
        ("TOTAL_REGISTROS_COMERCIALIZACAO", total_com),
        ("REGISTROS_ENCONTRADOS", encontrados),
        ("REGISTROS_APENAS_AGROFIT", apenas_agro),
        ("REGISTROS_APENAS_COMERCIALIZACAO", apenas_com),
        ("PERCENTUAL_CORRESPONDENCIA_REGISTROS", pct),
        ("VOLUME_PF_TOTAL", round(vol_total, 2)),
        ("VOLUME_PF_COM_MATCH_AGROFIT", round(vol_match, 2)),
        ("VOLUME_PF_SEM_AGROFIT", round(vol_sem, 2)),
        ("PERCENTUAL_VOLUME_COM_MATCH", pct_vol),
    ]
    return pd.DataFrame(linhas, columns=["INDICADOR", "VALOR"])


# =============================================================================
# RESUMO EXECUTIVO E LEGENDA
# =============================================================================

def montar_resumo_executivo(
    df_agrofit: pd.DataFrame,
    df_resumo_agrofit: pd.DataFrame,
    df_com_produtos: pd.DataFrame,
    df_com_uf: pd.DataFrame,
    df_vendas_ia: pd.DataFrame,
    df_validacao: Optional[pd.DataFrame] = None,
    df_evol_produtos: Optional[pd.DataFrame] = None,
) -> Tuple[pd.DataFrame, Dict[str, pd.DataFrame]]:
    indicadores = []

    def add(nome, valor):
        indicadores.append((nome, valor))

    add("Numero de ingredientes ativos (Agrofit)", df_agrofit["INGREDIENTE_ATIVO"].nunique() if "INGREDIENTE_ATIVO" in df_agrofit.columns else 0)
    add("Numero de produtos / marcas (Agrofit)", df_agrofit["MARCA_COMERCIAL"].nunique() if "MARCA_COMERCIAL" in df_agrofit.columns else 0)
    add("Numero de registros MAPA (Agrofit)", df_resumo_agrofit["_CHAVE_REGISTRO"].nunique() if "_CHAVE_REGISTRO" in df_resumo_agrofit.columns else 0)
    add("Numero de culturas (Agrofit)", df_agrofit["CULTURA"].nunique() if "CULTURA" in df_agrofit.columns else 0)
    add("Numero de pragas (Agrofit)", df_agrofit["PRAGA_NOME_CIENTIFICO"].nunique() if "PRAGA_NOME_CIENTIFICO" in df_agrofit.columns else 0)
    add("Numero de registros comercializados", df_com_produtos["_CHAVE_REGISTRO"].nunique() if not df_com_produtos.empty else 0)
    add(
        "Volume total comercializado PF (soma VENDAS_TOTAIS)",
        round(df_com_produtos["VENDAS_TOTAIS"].sum(), 2)
        if not df_com_produtos.empty and "VENDAS_TOTAIS" in df_com_produtos.columns else 0,
    )

    vol_ia = 0
    if not df_vendas_ia.empty:
        col_ia = next(
            (c for c in ["VENDA_PONDERADA_IA", "VENDA_ESTIMADA_IA", "VENDA_TOTAL"] if c in df_vendas_ia.columns),
            None,
        )
        if col_ia:
            vol_ia = round(df_vendas_ia[col_ia].sum(), 2)
    add("Volume estimado de IA comercializado (soma ponderada)", vol_ia)
    add("Numero de UFs com comercializacao", df_com_uf["UF"].nunique() if not df_com_uf.empty else 0)
    add(
        "Ano inicial",
        int(df_com_produtos["ANO"].min())
        if not df_com_produtos.empty and df_com_produtos["ANO"].notna().any() else None,
    )
    add(
        "Ano final",
        int(df_com_produtos["ANO"].max())
        if not df_com_produtos.empty and df_com_produtos["ANO"].notna().any() else None,
    )

    if df_validacao is not None and not df_validacao.empty:
        add("--- QUALIDADE DO CRUZAMENTO ---", "")
        for _, row in df_validacao.iterrows():
            add(str(row["INDICADOR"]), row["VALOR"])

    df_resumo = pd.DataFrame(indicadores, columns=["INDICADOR", "VALOR"])
    rankings: Dict[str, pd.DataFrame] = {}

    if not df_vendas_ia.empty and "VENDA_TOTAL" in df_vendas_ia.columns:
        top_ia = (
            df_vendas_ia.groupby("INGREDIENTE_ATIVO", as_index=False)["VENDA_TOTAL"].sum()
            .sort_values("VENDA_TOTAL", ascending=False).head(20).reset_index(drop=True)
        )
        rankings["Top20_Ingredientes_Ativos"] = top_ia

        anos = sorted(df_vendas_ia["ANO"].dropna().unique())
        if len(anos) >= 2:
            ano_ini, ano_fim = anos[0], anos[-1]
            base_ini = df_vendas_ia[df_vendas_ia["ANO"] == ano_ini].set_index("INGREDIENTE_ATIVO")["VENDA_TOTAL"]
            base_fim = df_vendas_ia[df_vendas_ia["ANO"] == ano_fim].set_index("INGREDIENTE_ATIVO")["VENDA_TOTAL"]
            cresc = pd.DataFrame({"VENDA_INICIAL": base_ini, "VENDA_FINAL": base_fim}).fillna(0)
            cresc = cresc[cresc["VENDA_INICIAL"] > 0]
            cresc["CRESCIMENTO_PERCENTUAL"] = round(
                100 * (cresc["VENDA_FINAL"] - cresc["VENDA_INICIAL"]) / cresc["VENDA_INICIAL"], 2
            )
            cresc = cresc.reset_index().sort_values("CRESCIMENTO_PERCENTUAL", ascending=False).head(10)
            rankings["Top10_IA_Crescimento"] = cresc

    if df_evol_produtos is not None and not df_evol_produtos.empty and "TOTAL_PERIODO" in df_evol_produtos.columns:
        cols_top = [c for c in ["NR_REGISTRO", "MARCA_COMERCIAL", "TOTAL_PERIODO"] if c in df_evol_produtos.columns]
        rankings["Top20_Produtos"] = (
            df_evol_produtos[cols_top]
            .sort_values("TOTAL_PERIODO", ascending=False)
            .head(20)
            .reset_index(drop=True)
        )
    elif not df_com_produtos.empty and "VENDAS_TOTAIS" in df_com_produtos.columns:
        top_prod = (
            df_com_produtos.groupby(["NR_REGISTRO", "MARCA_COMERCIAL"], as_index=False)["VENDAS_TOTAIS"]
            .sum().sort_values("VENDAS_TOTAIS", ascending=False).head(20).reset_index(drop=True)
        )
        rankings["Top20_Produtos"] = top_prod

    if not df_com_uf.empty:
        top_uf = (
            df_com_uf.groupby("UF", as_index=False)["VENDA"].sum()
            .sort_values("VENDA", ascending=False).head(10).reset_index(drop=True)
        )
        rankings["Top10_UF"] = top_uf

    return df_resumo, rankings


def montar_legenda() -> pd.DataFrame:
    linhas = [
        ("PREMISSA", "Perspectiva PRODUTO (PF)",
         "Volumes de produto formulado. Em multi-IA, o volume PF NAO e somado N vezes — usa-se first/max."),
        ("PREMISSA", "Perspectiva IA",
         "Cada ingrediente ativo permanece em linha propria; volume ponderado = volume PF x PONDERACAO_IA."),
        ("PREMISSA", "Colunas ORIGINAIS",
         "VENDAS_TOTAIS, VENDAS_CLIENTE, VENDAS_INDUSTRIA, PRODUCAO_NACIONAL, IMPORTACAO, EXPORTACAO, ESTOQUE_* — valores da planilha, sem ponderacao."),
        ("PREMISSA", "Colunas PONDERADAS",
         "VENDA_PONDERADA_IA / VENDA_ESTIMADA_IA e demais *_PONDERAD*_IA = original x teor de IA. Nunca sobrescrevem as originais."),
        ("PREMISSA", "Chave de cruzamento",
         "NR_REGISTRO normalizado (remove MAPA/IBAMA, pontuacao e zeros a esquerda) -> _CHAVE_REGISTRO."),
        ("PREMISSA", "Unidades",
         "Volumes conforme a base de comercializacao. Conferir arquivo IBAMA/MAPA de origem."),
        ("PREMISSA", "Regional",
         "Formato LONGO (Ano em linha). Multi-IA: max do volume PF por Registro x Ano x UF."),
        ("ABA", "00_Resumo", "KPIs, qualidade do match e rankings top."),
        ("ABA", "00_Legenda", "Este dicionario."),
        ("ABA", "01_Resumo_Agrofit",
         "1 linha por registro: cadastro Agrofit + NUM_CULTURAS/PRAGAS + principais culturas/pragas."),
        ("ABA", "02_Comercializacao_Produtos",
         "Fato PF: Registro x Ano (volumes originais de produto formulado + resumo de IAs)."),
        ("ABA", "03_Evolucao_Produtos",
         "WIDE produto x ano (volume PF) + TOTAL_PERIODO. Ideal para graficos de evolucao."),
        ("ABA", "04_Comercializacao_IA",
         "Fato IA: Registro x Ano x IA. Originais PF da linha + ponderados desta IA + STATUS_CALCULO_IA."),
        ("ABA", "05_Vendas_IA",
         "Agregacao IA x Ano (volumes ponderados de IA + contagens de produtos/registros)."),
        ("ABA", "06_Evolucao_IA",
         "WIDE IA x ano (volume ponderado) + TOTAL_PERIODO. Alias conceitual: Vendas_IA_Evolucao."),
        ("ABA", "07_Comerc_Regional_Produtos",
         "Vendas PF por Registro x Ano x UF (deduplicado com max em multi-IA)."),
        ("ABA", "08_Comerc_Regional_IA",
         "Estimativa de IA por UF = venda PF da UF x ponderacao da IA."),
        ("ABA", "09_Agrofit_Sem_Vendas", "Registros Agrofit sem comercializacao no periodo."),
        ("ABA", "10_Vendas_Sem_Agrofit", "Linhas de comercializacao sem correspondencia no Agrofit."),
        ("COLUNA", "VENDA_PONDERADA_IA",
         "Quantidade efetiva de ingrediente ativo (volume PF x teor). Alias: VENDA_ESTIMADA_IA."),
        ("COLUNA", "VENDA_TOTAL_PF_ASSOCIADA",
         "Soma de volumes PF das linhas daquele IA (NAO e volume de IA; rotulo explicito para evitar confusao)."),
        ("COLUNA", "STATUS_CALCULO_IA",
         "CALCULADO | SEM_PONDERACAO | UNIDADE_INCOMPATIVEL | DADO_AUSENTE."),
        ("COLUNA", "NUM_LINHAS_ORIGEM",
         "Quantidade de linhas da base bruta consolidadas no Registro x Ano (indica multi-IA)."),
        ("COLUNA", "TOTAL_PERIODO", "Soma dos volumes no horizonte de anos carregados."),
    ]
    return pd.DataFrame(linhas, columns=["TIPO", "ITEM", "DESCRICAO"])


# =============================================================================
# GERACAO DO EXCEL
# =============================================================================

ESTILO_CABECALHO_FONTE = Font(bold=True, color="FFFFFF", size=11)
ESTILO_CABECALHO_FUNDO = PatternFill("solid", fgColor="2E5B34")
ESTILO_AVISO_FONTE = Font(bold=True, color="9C4221", italic=True)
ESTILO_AVISO_FUNDO = PatternFill("solid", fgColor="FFF3CD")
LIMITE_LINHAS_TABELA_EXCEL = 900_000

# Caracteres de controle ilegais em XML/xlsx (exceto tab, LF, CR)
_RE_XML_ILLEGAL = re.compile(
    r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x84\x86-\x9F\uFDD0-\uFDEF"
    r"\uFFFE\uFFFF]"
)


def _sanitizar_valor_excel(val: Any) -> Any:
    """Converte valores para tipos aceitos pelo Excel/openpyxl, sem corromper o XML."""
    if val is None:
        return None
    if isinstance(val, (bool, np.bool_)):
        return bool(val)
    if isinstance(val, (np.integer,)):
        return int(val)
    if isinstance(val, (int,)):
        return val
    if isinstance(val, (np.floating, float)):
        f = float(val)
        if np.isnan(f) or np.isinf(f):
            return None
        return f
    if pd.isna(val):
        return None
    if isinstance(val, (pd.Timestamp, datetime)):
        try:
            if pd.isna(val):
                return None
            return val.to_pydatetime() if hasattr(val, "to_pydatetime") else val
        except Exception:
            return str(val)
    # Strings e demais: remove caracteres ilegais em XML
    s = str(val)
    if _RE_XML_ILLEGAL.search(s):
        s = _RE_XML_ILLEGAL.sub("", s)
    # Excel limita celulas a ~32k caracteres
    if len(s) > 32000:
        s = s[:31997] + "..."
    return s


def _escrever_dataframe_em_aba(
    ws, df: pd.DataFrame, aviso_topo: Optional[str] = None, nome_tabela: Optional[str] = None
):
    """Escreve um DataFrame em uma aba, de forma segura para o Excel abrir sem reparo."""
    linha_inicio = 1
    n_cols = max(len(df.columns), 1) if df is not None and not df.empty else 1

    if aviso_topo:
        cell = ws.cell(row=1, column=1, value=_sanitizar_valor_excel(aviso_topo))
        cell.font = ESTILO_AVISO_FONTE
        cell.fill = ESTILO_AVISO_FUNDO
        if n_cols > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        ws.row_dimensions[1].height = 30
        linha_inicio = 3

    if df is None or df.empty:
        ws.cell(row=linha_inicio, column=1, value="(sem dados para os arquivos fornecidos)")
        return

    df_out = df.copy()
    # Garante nomes de coluna unicos e strings limpas (Excel Table exige cabecalhos unicos)
    cols_limpos = []
    vistos = {}
    for c in df_out.columns:
        nome = _sanitizar_valor_excel(str(c)) or "COL"
        nome = str(nome).strip() or "COL"
        if nome in vistos:
            vistos[nome] += 1
            nome = f"{nome}_{vistos[nome]}"
        else:
            vistos[nome] = 0
        cols_limpos.append(nome)
    df_out.columns = cols_limpos

    if len(df_out) > LIMITE_LINHAS_TABELA_EXCEL:
        df_out = df_out.head(LIMITE_LINHAS_TABELA_EXCEL)

    for j, col in enumerate(df_out.columns, start=1):
        c = ws.cell(row=linha_inicio, column=j, value=str(col))
        c.font = ESTILO_CABECALHO_FONTE
        c.fill = ESTILO_CABECALHO_FUNDO
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for i, (_, row) in enumerate(df_out.iterrows(), start=linha_inicio + 1):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=_sanitizar_valor_excel(val))

    # freeze_panes como string (mais compativel entre versoes do openpyxl)
    ws.freeze_panes = f"A{linha_inicio + 1}"

    ultima_linha = linha_inicio + len(df_out)
    ultima_coluna = len(df_out.columns)
    # Tabela formatada so se houver ao menos 1 linha de dados alem do cabecalho
    if (
        nome_tabela
        and len(df_out) > 0
        and ultima_linha > linha_inicio
        and ultima_coluna >= 1
    ):
        try:
            # displayName: letra/underscore, sem espacos, unico no workbook
            nome_tab_limpo = re.sub(r"[^A-Za-z0-9_]", "", str(nome_tabela))
            if not nome_tab_limpo or nome_tab_limpo[0].isdigit():
                nome_tab_limpo = "T" + nome_tab_limpo
            ref = (
                f"{get_column_letter(1)}{linha_inicio}:"
                f"{get_column_letter(ultima_coluna)}{ultima_linha}"
            )
            tabela = Table(displayName=nome_tab_limpo, ref=ref)
            tabela.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False
            )
            ws.add_table(tabela)
        except Exception as e:
            LOGGER.warning("Nao foi possivel criar tabela formatada em %s: %s", nome_tabela, e)

    for j in range(1, ultima_coluna + 1):
        try:
            maior = max(
                [len(str(df_out.columns[j - 1]))]
                + [len(str(v)) for v in df_out.iloc[:, j - 1].head(100)]
            )
        except Exception:
            maior = 12
        ws.column_dimensions[get_column_letter(j)].width = min(max(maior + 2, 10), 40)


def gerar_excel(
    caminho_saida: Path,
    tabelas: Dict[str, pd.DataFrame],
    df_resumo: pd.DataFrame,
    rankings: Dict[str, pd.DataFrame],
    df_legenda: Optional[pd.DataFrame] = None,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # Nomes de aba <= 31 caracteres (limite do Excel). Ultrapassar corrompe o arquivo.
    # ---- 00_Resumo ----
    ws = wb.create_sheet("00_Resumo")
    _escrever_dataframe_em_aba(ws, df_resumo, nome_tabela="TabResumo")
    # Rankings abaixo do bloco de indicadores (fora da Table)
    linha_atual = (len(df_resumo) if df_resumo is not None and not df_resumo.empty else 0) + 4
    for nome_ranking, df_rank in rankings.items():
        if df_rank is None or df_rank.empty:
            continue
        ws.cell(row=linha_atual, column=1, value=_sanitizar_valor_excel(nome_ranking.replace("_", " "))).font = Font(
            bold=True, size=12
        )
        linha_atual += 1
        for j, col in enumerate(df_rank.columns, start=1):
            c = ws.cell(row=linha_atual, column=j, value=_sanitizar_valor_excel(str(col)))
            c.font = ESTILO_CABECALHO_FONTE
            c.fill = ESTILO_CABECALHO_FUNDO
        linha_atual += 1
        for _, row in df_rank.iterrows():
            for j, val in enumerate(row, start=1):
                ws.cell(row=linha_atual, column=j, value=_sanitizar_valor_excel(val))
            linha_atual += 1
        linha_atual += 2

    # ---- 00_Legenda ----
    ws = wb.create_sheet("00_Legenda")
    _escrever_dataframe_em_aba(
        ws,
        df_legenda if df_legenda is not None else montar_legenda(),
        nome_tabela="TabLegenda",
    )

    fill_original = PatternFill("solid", fgColor="D6EAF8")
    fill_ponderado = PatternFill("solid", fgColor="D5F5E3")
    cols_originais = {
        "ESTOQUE_INICIAL", "PRODUCAO_NACIONAL", "IMPORTACAO", "EXPORTACAO",
        "VENDAS_INDUSTRIA", "VENDAS_CLIENTE", "VENDAS_TOTAIS", "ESTOQUE_FINAL",
    }
    cols_ponderadas = {
        "ESTOQUE_INICIAL_PONDERADO_IA", "PRODUCAO_NACIONAL_PONDERADA_IA",
        "IMPORTACAO_PONDERADA_IA", "EXPORTACAO_PONDERADA_IA",
        "VENDAS_INDUSTRIA_PONDERADA_IA", "VENDAS_CLIENTE_PONDERADA_IA",
        "VENDA_PONDERADA_IA", "VENDA_ESTIMADA_IA", "ESTOQUE_FINAL_PONDERADO_IA",
        "PONDERACAO_IA",
    }

    # nome_aba deve ter no maximo 31 caracteres
    abas_numeradas = [
        ("01_Resumo_Agrofit", "Resumo_Agrofit", "TabResumoAgrofit"),
        ("02_Comercializacao_Produtos", "Comercializacao_Produtos", "TabComercProdutos"),
        ("03_Evolucao_Produtos", "Evolucao_Produtos", "TabEvolProdutos"),
        ("04_Comercializacao_IA", "Comercializacao_IA", "TabComercIA"),
        ("05_Vendas_IA", "Vendas_IA", "TabVendasIA"),
        ("06_Evolucao_IA", "Evolucao_IA", "TabEvolIA"),
        ("07_Comerc_Regional_Produtos", "Comercializacao_Regional_Produtos", "TabRegProd"),
        ("08_Comerc_Regional_IA", "Comercializacao_Regional_IA", "TabRegIA"),
        ("09_Agrofit_Sem_Vendas", "Agrofit_Sem_Vendas", "TabAgrofitSemVendas"),
        ("10_Vendas_Sem_Agrofit", "Vendas_Sem_Agrofit", "TabVendasSemAgrofit"),
    ]

    for nome_aba, chave_tabela, nome_tab in abas_numeradas:
        if len(nome_aba) > 31:
            raise ValueError(f"Nome de aba excede 31 caracteres: {nome_aba!r} ({len(nome_aba)})")
        ws = wb.create_sheet(nome_aba)
        aviso_txt = None
        if nome_aba == "04_Comercializacao_IA":
            aviso_txt = (
                "Azul claro = volumes ORIGINAIS (PF da linha). "
                "Verde claro = volumes PONDERADOS pela IA (teor x PF). "
                "Nao some volumes PF entre linhas da mesma formulacao multi-IA."
            )
        df_aba = tabelas.get(chave_tabela, pd.DataFrame())
        _escrever_dataframe_em_aba(ws, df_aba, aviso_topo=aviso_txt, nome_tabela=nome_tab)

        if nome_aba == "04_Comercializacao_IA" and df_aba is not None and not df_aba.empty:
            header_row = 3 if aviso_txt else 1
            for j, col in enumerate(df_aba.columns, start=1):
                cell = ws.cell(row=header_row, column=j)
                if str(col) in cols_originais:
                    cell.fill = fill_original
                    cell.font = Font(bold=True, color="1A5276")
                elif str(col) in cols_ponderadas:
                    cell.fill = fill_ponderado
                    cell.font = Font(bold=True, color="145A32")

    caminho_saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho_saida)
    wb.close()



# =============================================================================
# ORQUESTRACAO
# =============================================================================

def preparar_pastas() -> None:
    for pasta in [DADOS_DIR, SAIDA_DIR, LOGS_DIR]:
        pasta.mkdir(parents=True, exist_ok=True)


def main() -> None:
    tempo_total_inicio = time.time()
    preparar_pastas()
    configurar_logging()

    titulo("ANALISE INTEGRADA AGROFIT x COMERCIALIZACAO DE AGROTOXICOS")
    LOGGER.info("Execucao iniciada em %s", datetime.now().isoformat())

    # 1) Agrofit
    caminho_agrofit = localizar_arquivo_agrofit()
    if caminho_agrofit is None:
        erro(
            f"Arquivo do Agrofit nao encontrado em {DADOS_DIR}. "
            f"Coloque '{ARQUIVO_AGROFIT_PADRAO}' na pasta 'dados/' e execute novamente."
        )
        if os.environ.get("AGROFIT_NO_INPUT"):
            return
        input("\nPressione Enter para sair...")
        return

    with Cronometro(f"Carregar base Agrofit ({caminho_agrofit.name})"):
        df_agrofit = carregar_agrofit(caminho_agrofit)

    if df_agrofit.empty:
        erro("Base Agrofit vazia ou sem registros validos. Encerrando.")
        if os.environ.get("AGROFIT_NO_INPUT"):
            return
        input("\nPressione Enter para sair...")
        return

    # 2) Comercializacao
    arquivos_com = detectar_arquivos_comercializacao(caminho_agrofit)
    if not arquivos_com:
        aviso(
            f"Nenhum arquivo de comercializacao em {DADOS_DIR}. "
            "Relatorio apenas com informacoes regulatorias do Agrofit."
        )
    else:
        print(f"\n{len(arquivos_com)} arquivo(s) de comercializacao detectado(s):")
        for a in arquivos_com:
            print(f"   - {a.name}")

    with Cronometro("Carregar bases de comercializacao"):
        df_vendas_bruto = carregar_todas_comercializacoes(arquivos_com)

    if not df_vendas_bruto.empty:
        with Cronometro("Preparar campos numericos de comercializacao"):
            df_vendas_bruto = preparar_numericos_comercializacao(df_vendas_bruto)
        if "INGREDIENTE_ATIVO" in df_vendas_bruto.columns:
            df_vendas_bruto["INGREDIENTE_ATIVO"] = df_vendas_bruto["INGREDIENTE_ATIVO"].apply(
                lambda v: v if pd.isna(v) else str(v).strip()
            )

    # 3) Tabelas analiticas
    with Cronometro("01_Resumo_Agrofit"):
        df_resumo_agrofit = montar_resumo_agrofit(df_agrofit)
        ok(f"{len(df_resumo_agrofit):,} registros unicos no Agrofit")

    with Cronometro("02_Comercializacao_Produtos"):
        df_com_produtos = montar_comercializacao_produtos(df_vendas_bruto, df_resumo_agrofit)
        ok(f"{len(df_com_produtos):,} linhas (produto x ano)")

    with Cronometro("03_Evolucao_Produtos"):
        df_evol_produtos = montar_evolucao_produtos(df_com_produtos)
        ok(f"{len(df_evol_produtos):,} produtos com serie historica")

    with Cronometro("04_Comercializacao_IA"):
        df_com_ia = montar_comercializacao_ia(df_vendas_bruto, df_resumo_agrofit)
        ok(f"{len(df_com_ia):,} linhas (registro x ano x IA)")

    with Cronometro("05_Vendas_IA"):
        df_vendas_ia = montar_vendas_ia(df_com_ia)
        ok(f"{len(df_vendas_ia):,} linhas (IA x ano)")

    with Cronometro("06_Evolucao_IA"):
        df_evol_ia = montar_evolucao_ia(df_vendas_ia)
        ok(f"{len(df_evol_ia):,} ingredientes ativos com serie historica")

    with Cronometro("07_Comercializacao_Regional_Produtos"):
        df_com_uf = montar_comercializacao_regional_produtos(df_vendas_bruto, df_resumo_agrofit)
        ok(f"{len(df_com_uf):,} linhas (registro x ano x UF)")

    with Cronometro("08_Comercializacao_Regional_IA"):
        df_ia_uf = montar_comercializacao_regional_ia(df_com_uf, df_com_ia)
        ok(f"{len(df_ia_uf):,} linhas (IA x ano x UF)")

    with Cronometro("09_Agrofit_Sem_Vendas"):
        df_agrofit_sem = montar_agrofit_sem_vendas(df_resumo_agrofit, df_com_produtos)
        ok(f"{len(df_agrofit_sem):,} registros sem comercializacao")

    with Cronometro("10_Vendas_Sem_Agrofit"):
        df_vendas_sem = montar_vendas_sem_agrofit(df_com_produtos, df_resumo_agrofit)
        if len(df_vendas_sem) > 0:
            aviso(f"{len(df_vendas_sem):,} registros de comercializacao sem match no Agrofit")

    with Cronometro("Validacao e resumo executivo"):
        df_validacao = montar_validacao(
            df_resumo_agrofit, df_com_produtos, df_agrofit_sem, df_vendas_sem
        )
        df_resumo, rankings = montar_resumo_executivo(
            df_agrofit, df_resumo_agrofit, df_com_produtos, df_com_uf, df_vendas_ia,
            df_validacao=df_validacao,
            df_evol_produtos=df_evol_produtos,
        )
        df_legenda = montar_legenda()

    # Flags de cobertura no Resumo_Agrofit
    if not df_resumo_agrofit.empty:
        chaves_com = set(df_com_produtos["_CHAVE_REGISTRO"].unique()) if not df_com_produtos.empty else set()
        df_resumo_agrofit = df_resumo_agrofit.copy()
        df_resumo_agrofit["TEM_COMERCIALIZACAO"] = df_resumo_agrofit["_CHAVE_REGISTRO"].apply(
            lambda k: "SIM" if k in chaves_com else "NAO"
        )

    tabelas = {
        "Resumo_Agrofit": df_resumo_agrofit.drop(columns=["_CHAVE_REGISTRO"], errors="ignore"),
        "Comercializacao_Produtos": df_com_produtos.drop(columns=["_CHAVE_REGISTRO"], errors="ignore"),
        "Evolucao_Produtos": df_evol_produtos,
        "Comercializacao_IA": df_com_ia.drop(columns=["_CHAVE_REGISTRO"], errors="ignore"),
        "Vendas_IA": df_vendas_ia,
        "Evolucao_IA": df_evol_ia,
        "Comercializacao_Regional_Produtos": df_com_uf.drop(columns=["_CHAVE_REGISTRO"], errors="ignore"),
        "Comercializacao_Regional_IA": df_ia_uf,
        "Agrofit_Sem_Vendas": df_agrofit_sem,
        "Vendas_Sem_Agrofit": df_vendas_sem,
    }

    # 4) Excel
    caminho_saida = SAIDA_DIR / ARQUIVO_SAIDA
    with Cronometro(f"Gerar arquivo Excel final ({ARQUIVO_SAIDA})"):
        gerar_excel(caminho_saida, tabelas, df_resumo, rankings, df_legenda=df_legenda)

    tempo_total = time.time() - tempo_total_inicio
    LOGGER.info("Execucao concluida em %.2fs", tempo_total)

    titulo("PROCESSAMENTO CONCLUIDO")
    print(f"Arquivo gerado : {caminho_saida}")
    print(f"Log detalhado  : {LOGS_DIR / ARQUIVO_LOG}")
    print(f"Tempo total    : {tempo_legivel(tempo_total)}\n")

    if not df_validacao.empty:
        pct = df_validacao.loc[
            df_validacao["INDICADOR"] == "PERCENTUAL_CORRESPONDENCIA_REGISTROS", "VALOR"
        ]
        if len(pct):
            print(f"Percentual de correspondencia (registros): {pct.iloc[0]}%")
        pct_vol = df_validacao.loc[
            df_validacao["INDICADOR"] == "PERCENTUAL_VOLUME_COM_MATCH", "VALOR"
        ]
        if len(pct_vol):
            print(f"Percentual de correspondencia (volume PF): {pct_vol.iloc[0]}%")
    print()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        aviso("Interrompido pelo usuario.")
    except Exception as e:
        erro(f"Erro inesperado: {e}")
        LOGGER.error("Erro inesperado: %s\n%s", e, traceback.format_exc())
        traceback.print_exc()
        if not os.environ.get("AGROFIT_NO_INPUT"):
            input("\nPressione Enter para sair...")
