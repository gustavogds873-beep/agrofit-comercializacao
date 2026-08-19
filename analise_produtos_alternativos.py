#!/usr/bin/env python
# -*- coding: utf-8 -*-
#  ANÁLISE AGROFIT — PRODUTOS ALTERNATIVOS
# ================================================================
#  Script completo e standalone para Windows / Linux / macOS / Colab
#
#  Funcionalidades (análises exclusivas da base Agrofit):
#    1  Busca por CULTURA          → resumo pragas × IAs
#    2  Busca por PRAGA            → resumo culturas × IAs
#    3  Busca por IA + Alternativos → Excel/CSV de produtos alternativos
#       - Organizados por CULTURA (abas por praga) — comportamento original
#       - Organizados por PRAGA   (abas Resumo_IA / Relatorio / Produtos_Alternativos)
#    4  Produtos que CONTÊM o IA   → tabela para preenchimento de dose
#
#  Análises que cruzam Agrofit com comercialização ficam no script
#  especializado: analise_integrada_agrofit_comercializacao.py
#
#  Requisitos (instalar uma vez):
#    pip install pandas requests openpyxl unidecode rapidfuzz tqdm
#
#  Uso local:
#    python analise_produtos_alternativos.py
#
#  Uso no Google Colab:
#    - Defina AGROFIT_DADOS_DIR, AGROFIT_SAIDA_DIR (opcional AGROFIT_NO_INPUT=1)
#    - Ou use o notebook Analise_Produtos_Alternativos.ipynb
#
#  Arquivos esperados:
#    - agrofitprodutosformulados.csv  (baixado automaticamente se ausente)
#
#  Repositório: https://github.com/gustavogds873-beep/agrofit-comercializacao
#  Autor base: Gustavo Gomes de Sousa
# ================================================================

from __future__ import annotations

import os
import sys
import re
import time
import glob
import warnings
from pathlib import Path
from collections import Counter
from datetime import datetime
from typing import Optional, Union, List, Dict, Tuple, Any

# ---------------------------------------------------------------------------
# Dependências obrigatórias / opcionais
# ---------------------------------------------------------------------------

def _ensure_packages():
    """Instala pacotes faltantes de forma silenciosa (primeira execução)."""
    required = ["pandas", "requests", "openpyxl", "unidecode", "rapidfuzz", "tqdm"]
    missing = []
    for pkg in required:
        try:
            __import__(pkg if pkg != "openpyxl" else "openpyxl")
        except ImportError:
            missing.append(pkg)
    if missing:
        print(f"Instalando pacotes faltantes: {', '.join(missing)} ...")
        import subprocess
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "-q"] + missing,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        print("Pacotes instalados.\n")


_ensure_packages()

import pandas as pd
import numpy as np
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

try:
    from tqdm import tqdm
    TQDM_AVAILABLE = True
except ImportError:
    TQDM_AVAILABLE = False

try:
    import unidecode as _unidecode_mod
    from unidecode import unidecode
    UNIDECODE_AVAILABLE = True
except ImportError:
    UNIDECODE_AVAILABLE = False

    def unidecode(s):  # fallback
        return s

try:
    from rapidfuzz import fuzz, process
    RAPIDFUZZ_AVAILABLE = True
except ImportError:
    RAPIDFUZZ_AVAILABLE = False

warnings.filterwarnings("ignore", category=UserWarning)
warnings.filterwarnings("ignore", category=FutureWarning)

# ---------------------------------------------------------------------------
# Cores ANSI (funcionam no Windows 10+ com VT habilitado)
# ---------------------------------------------------------------------------

class bcolors:
    HEADER = "\033[95m"
    OKBLUE = "\033[94m"
    OKGREEN = "\033[92m"
    WARNING = "\033[93m"
    FAIL = "\033[91m"
    ENDC = "\033[0m"
    BOLD = "\033[1m"


# Habilita ANSI no Windows
if sys.platform == "win32":
    try:
        import ctypes
        kernel32 = ctypes.windll.kernel32
        kernel32.SetConsoleMode(kernel32.GetStdHandle(-11), 7)
    except Exception:
        bcolors.HEADER = bcolors.OKBLUE = bcolors.OKGREEN = ""
        bcolors.WARNING = bcolors.FAIL = bcolors.ENDC = bcolors.BOLD = ""


def print_header(text: str):
    print(f"\n{bcolors.HEADER}{bcolors.BOLD}{'═' * 70}{bcolors.ENDC}")
    print(f"{bcolors.BOLD}{text.center(70)}{bcolors.ENDC}")
    print(f"{bcolors.HEADER}{'═' * 70}{bcolors.ENDC}\n")


def print_success(msg: str):
    print(f"{bcolors.OKGREEN}✓ {msg}{bcolors.ENDC}")


def print_warning(msg: str):
    print(f"{bcolors.WARNING}⚠ {msg}{bcolors.ENDC}")


def print_error(msg: str):
    print(f"{bcolors.FAIL}✗ {msg}{bcolors.ENDC}")


def segundos_para_tempo_legivel(segundos: float) -> str:
    if segundos < 60:
        return f"{int(segundos)}s"
    if segundos < 3600:
        return f"{segundos / 60:.1f} min"
    if segundos < 86400:
        return f"{segundos / 3600:.1f} h"
    return f"{segundos / 86400:.1f} dias"


def sanitize_filename(name: str) -> str:
    name = re.sub(r'[<>:"/\\|?*]', "_", str(name))
    name = re.sub(r"\s+", "_", name.strip())
    return name[:200]


def extrair_apenas_numeros(valor) -> str:
    if pd.isna(valor):
        return ""
    return re.sub(r"\D", "", str(valor).strip())


def extrair_chave_flexivel(valor) -> str:
    """MAPA-03009 / 03009 / 3009 → '3009'."""
    if pd.isna(valor):
        return ""
    s = str(valor).strip().upper()
    s = re.sub(r"^(MAPA|IBAMA)[-/]?", "", s)
    nums = re.sub(r"\D", "", s)
    return nums.lstrip("0") if nums else ""


def smart_progress(iterable, total=None, desc="Processando", unit="it"):
    if TQDM_AVAILABLE:
        return tqdm(
            iterable,
            total=total,
            desc=desc,
            unit=unit,
            bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}]",
            dynamic_ncols=True,
            miniters=1,
        )

    class FakeProgress:
        def __init__(self, total):
            self.total = total
            self.count = 0

        def update(self, n=1):
            self.count += n
            if self.total and (self.count % max(1, self.total // 20) == 0 or self.count == self.total):
                perc = self.count / self.total * 100
                print(f"  {desc}: {self.count:,}/{self.total:,} ({perc:5.1f}%)")

        def close(self):
            if self.total:
                print(f"  {desc} concluído: {self.count:,}/{self.total:,}")

        def __iter__(self):
            for item in iterable:
                yield item
                self.update(1)
            self.close()

    return FakeProgress(total or (len(iterable) if hasattr(iterable, "__len__") else None))


# ---------------------------------------------------------------------------
# Configuração global
# ---------------------------------------------------------------------------

# Pastas: suportam execução local e Google Colab via variáveis de ambiente
#   AGROFIT_DADOS_DIR  → onde está / será baixado o CSV Agrofit
#   AGROFIT_SAIDA_DIR  → onde gravar os resultados
#   AGROFIT_LOGS_DIR   → (reservado)
#   AGROFIT_NO_INPUT   → "1" força respostas padrão em prompts de download
#
# Prioridade de resolução:
#   1. Variáveis de ambiente (AGROFIT_*)
#   2. Detecção automática do Colab (/content/dados e /content/resultados)
#   3. Pasta do script (execução local)

def _detectar_colab() -> bool:
    """Retorna True se estiver rodando no Google Colab."""
    if Path("/content").is_dir():
        try:
            import google.colab  # noqa: F401
            return True
        except ImportError:
            # Pasta /content existe (ambiente tipo Colab) mesmo sem o pacote
            return Path("/content/dados").exists() or Path("/content/scripts").exists()
    return False


def _resolver_pastas():
    """
    Resolve DADOS_DIR, SAIDA_DIR, LOGS_DIR e CSV_LOCAL.
    Pode ser chamada de novo no início de main() para pegar env vars
    definidas logo antes do %run no notebook.
    """
    try:
        script_file = Path(__file__).resolve()
        script_dir = script_file.parent
    except NameError:
        script_dir = Path.cwd()

    em_colab = _detectar_colab()

    # Defaults: no Colab usa /content/{dados,resultados,logs}; local usa pasta do script
    if em_colab:
        default_dados = Path("/content/dados")
        default_saida = Path("/content/resultados")
        default_logs = Path("/content/logs")
    else:
        default_dados = script_dir
        default_saida = script_dir
        default_logs = script_dir / "logs"

    dados = Path(os.environ.get("AGROFIT_DADOS_DIR", str(default_dados))).resolve()
    saida = Path(os.environ.get("AGROFIT_SAIDA_DIR", str(default_saida))).resolve()
    logs = Path(os.environ.get("AGROFIT_LOGS_DIR", str(default_logs))).resolve()
    no_input = os.environ.get("AGROFIT_NO_INPUT", "").strip().lower() in (
        "1", "true", "yes", "y", "sim"
    )

    dados.mkdir(parents=True, exist_ok=True)
    saida.mkdir(parents=True, exist_ok=True)
    logs.mkdir(parents=True, exist_ok=True)

    csv_local = dados / "agrofitprodutosformulados.csv"

    # Se o CSV ainda não está em dados/, procura em locais comuns do Colab
    if not csv_local.exists():
        candidatos = [
            Path("/content/dados/agrofitprodutosformulados.csv"),
            Path("/content/agrofitprodutosformulados.csv"),
            script_dir / "agrofitprodutosformulados.csv",
            Path.cwd() / "agrofitprodutosformulados.csv",
        ]
        for c in candidatos:
            if c.exists() and c.stat().st_size > 1_000_000:
                # Usa o caminho encontrado (não copia; apenas aponta)
                csv_local = c
                break

    return script_dir, dados, saida, logs, no_input, str(csv_local)


SCRIPT_DIR, DADOS_DIR, SAIDA_DIR, LOGS_DIR, NO_INPUT, CSV_LOCAL = _resolver_pastas()

AGROFIT_CSV_URL = (
    "https://dados.agricultura.gov.br/dataset/"
    "6c913699-e82e-4da3-a0a1-fb6c431e367f/resource/"
    "d30b30d7-e256-484e-9ab8-cd40974e1238/download/agrofitprodutosformulados.csv"
)
SEP = ";"

TOX_MAP = {
    "extremamente tóxico": 1, "categoria 1": 1,
    "altamente tóxico": 2, "categoria 2": 2,
    "medianamente tóxico": 3, "moderadamente tóxico": 3, "categoria 3": 3,
    "pouco tóxico": 4, "categoria 4": 4,
    "categoria 5": 5, "improvável de causar dano agudo": 5,
    "não classificado": 99, "não determinado": 99, "baixa exposição": 99,
}

AMB_MAP = {
    "altamente perigoso ao meio ambiente": 1,
    "muito perigoso ao meio ambiente": 2,
    "perigoso ao meio ambiente": 3,
    "pouco perigoso ao meio ambiente": 4,
    "baixo risco ao meio ambiente": 5,
    "não classificado": 99,
}

UFS = [
    "RO", "AC", "AM", "RR", "PA", "AP", "TO", "MA", "PI", "CE", "RN", "PB",
    "PE", "AL", "SE", "BA", "MG", "ES", "RJ", "SP", "PR", "SC", "RS", "MS",
    "MT", "GO", "DF",
]

REGIOES = {
    "Norte": ["RO", "AC", "AM", "RR", "PA", "AP", "TO"],
    "Nordeste": ["MA", "PI", "CE", "RN", "PB", "PE", "AL", "SE", "BA"],
    "Centro-Oeste": ["MT", "MS", "GO", "DF"],
    "Sudeste": ["MG", "ES", "RJ", "SP"],
    "Sul": ["PR", "SC", "RS"],
}


# ================================================================
# BUSCA INTERATIVA COM FUZZY + SELEÇÃO
# ================================================================

def busca_interativa_e_selecao(
    df: pd.DataFrame, coluna: str, termo: str, max_sugestoes: int = 20
) -> list:
    """Busca interativa com fuzzy matching e seleção múltipla pelo usuário."""
    if not termo.strip():
        print_error("Termo de busca vazio.")
        return []

    termo_norm = termo.lower().strip()
    usa_unidecode = UNIDECODE_AVAILABLE
    termo_norm_unaccent = unidecode(termo_norm) if usa_unidecode else termo_norm

    print(f"\n{bcolors.OKBLUE}Buscando por '{termo}' na coluna '{coluna}'...{bcolors.ENDC}")

    valores_unicos = df[coluna].dropna().astype(str).str.strip().unique()
    if len(valores_unicos) == 0:
        print_warning("Nenhum valor não-nulo encontrado.")
        return []

    if usa_unidecode:
        valores_norm = [unidecode(v.lower().strip()) for v in valores_unicos]
    else:
        valores_norm = [v.lower().strip() for v in valores_unicos]

    freq = Counter(df[coluna].astype(str).str.strip())
    candidatos = []

    if RAPIDFUZZ_AVAILABLE:
        if coluna in ["CULTURA", "PRAGA_NOME_CIENTIFICO", "PRAGA_NOME_COMUM"]:
            scorer = fuzz.WRatio
            threshold = 75
            compr_max_fator = 3
        else:
            scorer = fuzz.token_sort_ratio
            threshold = 85
            compr_max_fator = 2.5

        print(f"→ Usando rapidfuzz ({scorer.__name__}, threshold={threshold})...")
        matches = process.extract(
            termo_norm_unaccent if usa_unidecode else termo_norm,
            valores_norm,
            scorer=scorer,
            limit=max_sugestoes * 3,
        )
        for match, score, idx in matches:
            nome_original = valores_unicos[idx]
            if score >= threshold and len(nome_original) <= len(termo_norm) * compr_max_fator:
                candidatos.append((nome_original, score))
    else:
        print_warning("rapidfuzz não encontrado → busca parcial simples")
        mask = pd.Series(valores_norm).str.contains(
            termo_norm_unaccent if usa_unidecode else termo_norm, na=False
        )
        candidatos = [(valores_unicos[i], 100) for i in range(len(valores_unicos)) if mask[i]]

    candidatos_ordenados = sorted(
        candidatos, key=lambda x: (-x[1], -freq[x[0]])
    )[:max_sugestoes]

    if not candidatos_ordenados:
        print_warning("Nenhuma sugestão fuzzy → fallback 'contém'")
        mask_contains = df[coluna].astype(str).str.lower().str.contains(termo_norm, na=False)
        if mask_contains.any():
            variacoes = df[mask_contains][coluna].value_counts().index.tolist()
            candidatos_ordenados = [(v, 100) for v in variacoes[:max_sugestoes]]

    if not candidatos_ordenados:
        print_warning(f"Nenhuma sugestão encontrada para '{termo}'")
        return []

    print_success(f"Encontradas {len(candidatos_ordenados)} sugestões relevantes:")
    print(f"{'Nr':>3} | {'Nome encontrado':<60} | {'Ocorrências':>12} | {'Score':>6}")
    print("-" * 90)
    for i, (nome, score) in enumerate(candidatos_ordenados, 1):
        print(f"{i:>3} | {nome:<60} | {freq[nome]:>12,} | {score:>6.0f}")

    print("\nDigite os números desejados (ex: 1, 3-5, 8, todos, nenhum, cancelar):")
    while True:
        selecao = input("> ").strip().lower()
        if selecao in ["cancelar", "c", "voltar", "v"]:
            return []
        if selecao in ["nenhum", "n"]:
            return []
        if selecao in ["todos", "t", "all", "a"]:
            return [nome for nome, _ in candidatos_ordenados]

        selecionados = []
        try:
            partes = selecao.replace(" ", "").split(",")
            for parte in partes:
                if "-" in parte:
                    ini, fim = map(int, parte.split("-"))
                    selecionados.extend(range(ini, fim + 1))
                else:
                    selecionados.append(int(parte))
            selecionados = sorted(set(selecionados))
            valores_escolhidos = []
            for num in selecionados:
                if 1 <= num <= len(candidatos_ordenados):
                    valores_escolhidos.append(candidatos_ordenados[num - 1][0])
                else:
                    print_warning(f"Número {num} inválido → ignorado")
            if valores_escolhidos:
                print_success(f"Selecionados {len(valores_escolhidos)} itens:")
                for v in valores_escolhidos:
                    print(f"  • {v}")
                return valores_escolhidos
            print_warning("Nenhum número válido. Tente novamente.")
        except Exception:
            print_warning("Formato inválido. Exemplos: 1  ou  3-7  ou  1,4,8  ou  todos")


# ================================================================
# CARREGAMENTO DA BASE AGROFIT
# ================================================================

def carregar_dados_agrofit(forcar_download: bool = False) -> Optional[pd.DataFrame]:
    print_header("CARREGAMENTO DA BASE DE DADOS AGROFIT")

    arquivo_existe = os.path.exists(CSV_LOCAL)
    deve_baixar = forcar_download or not arquivo_existe

    if arquivo_existe and not forcar_download:
        tempo_modif = os.path.getmtime(CSV_LOCAL)
        dt_modif = datetime.fromtimestamp(tempo_modif)
        idade = time.time() - tempo_modif
        tamanho = os.path.getsize(CSV_LOCAL)
        print(f"→ Arquivo local: {CSV_LOCAL}")
        print(f"  Última atualização: {dt_modif.strftime('%d/%m/%Y %H:%M:%S')}")
        print(f"  Idade: {segundos_para_tempo_legivel(idade)}")
        print(f"  Tamanho: {tamanho / (1024 * 1024):.1f} MB")
        if NO_INPUT:
            # No Colab / modo não-interativo: reutiliza o arquivo local
            deve_baixar = False
            print_success("Modo não-interativo: reutilizando arquivo local.")
        else:
            print("\nDeseja baixar a versão mais recente? (s/n)")
            while True:
                resp = input("> ").strip().lower()
                if resp in ["s", "sim", "y", "yes"]:
                    deve_baixar = True
                    break
                if resp in ["n", "nao", "não", "no"]:
                    deve_baixar = False
                    break
                print_warning("Digite 's' ou 'n'.")

    if deve_baixar:
        print("\nIniciando download da base Agrofit...")
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        }
        total_esperado = 0
        try:
            r_head = requests.head(AGROFIT_CSV_URL, headers=headers, timeout=20, allow_redirects=True)
            if r_head.status_code in (200, 206):
                total_esperado = int(r_head.headers.get("Content-Length", 0))
                if total_esperado:
                    print(f"Tamanho esperado: {total_esperado / (1024 * 1024):.1f} MB")
        except Exception:
            pass

        max_tentativas = 5
        for tentativa in range(1, max_tentativas + 1):
            try:
                with requests.get(
                    AGROFIT_CSV_URL, headers=headers, stream=True, timeout=(30, 1800)
                ) as r:
                    r.raise_for_status()
                    with open(CSV_LOCAL, "wb") as f:
                        if TQDM_AVAILABLE and total_esperado:
                            pbar = tqdm(
                                total=total_esperado, unit="B", unit_scale=True,
                                desc="Baixando Agrofit",
                            )
                        else:
                            pbar = None
                        for chunk in r.iter_content(chunk_size=128 * 1024):
                            if chunk:
                                f.write(chunk)
                                if pbar:
                                    pbar.update(len(chunk))
                        if pbar:
                            pbar.close()
                tamanho_final = os.path.getsize(CSV_LOCAL)
                print_success(f"Download finalizado ({tamanho_final / (1024 * 1024):.1f} MB)")
                break
            except Exception as e:
                print_error(f"Erro na tentativa {tentativa}/{max_tentativas}: {e}")
                if tentativa == max_tentativas:
                    if not arquivo_existe:
                        return None
                    print_warning("Usando arquivo local existente.")
                time.sleep(5 * tentativa)

    print("\nCarregando base de dados...")
    try:
        df = pd.read_csv(CSV_LOCAL, sep=SEP, low_memory=False, dtype=str)
        print_success(f"Base carregada: {len(df):,} registros")
    except Exception as e:
        print_error(f"Erro ao carregar CSV: {e}")
        return None

    colunas_esperadas = [
        "CULTURA", "PRAGA_NOME_CIENTIFICO", "PRAGA_NOME_COMUM", "INGREDIENTE_ATIVO"
    ]
    faltando = [c for c in colunas_esperadas if c not in df.columns]
    if faltando:
        print_error(f"Colunas obrigatórias ausentes: {faltando}")
        return None

    df["CHAVE_FLEXIVEL"] = df["NR_REGISTRO"].apply(extrair_chave_flexivel) if "NR_REGISTRO" in df.columns else ""
    return df


# ================================================================
# MODOS 1–4 (resumos e alternativos)
# ================================================================

def gerar_resumo_por_cultura(df: pd.DataFrame, cultura: str):
    print(f"\n{bcolors.OKBLUE}MODO 1 - Busca por CULTURA → {cultura}{bcolors.ENDC}")
    if not cultura.strip():
        return None, "Cultura inválida"
    culturas_selecionadas = busca_interativa_e_selecao(df, "CULTURA", cultura, max_sugestoes=15)
    if not culturas_selecionadas:
        return None, "Cancelado pelo usuário"

    df_filtrado = df[df["CULTURA"].isin(culturas_selecionadas)].copy()
    print_success(f"Registros filtrados: {len(df_filtrado):,}")
    df_filtrado["PRAGA_NOME_COMUM"] = df_filtrado["PRAGA_NOME_COMUM"].fillna("Não informado")

    resumo = (
        df_filtrado.groupby(
            ["PRAGA_NOME_CIENTIFICO", "PRAGA_NOME_COMUM", "INGREDIENTE_ATIVO"],
            as_index=False,
        )
        .size()
        .rename(columns={"size": "Nº de Produtos Registrados"})
    )
    resumo = resumo.sort_values(
        ["Nº de Produtos Registrados", "PRAGA_NOME_CIENTIFICO"], ascending=[False, True]
    ).reset_index(drop=True)
    resumo.index += 1
    resumo = resumo.rename(columns={
        "PRAGA_NOME_CIENTIFICO": "Praga (Nome Científico)",
        "PRAGA_NOME_COMUM": "Praga (Nome Comum)",
        "INGREDIENTE_ATIVO": "Ingrediente Ativo",
    })
    print_success(f"Resumo gerado com {len(resumo):,} combinações")
    return resumo, None


def gerar_resumo_por_praga(df: pd.DataFrame, praga: str):
    print(f"\n{bcolors.OKBLUE}MODO 2 - Busca por PRAGA → {praga}{bcolors.ENDC}")
    if not praga.strip():
        return None, "Praga inválida"

    print("Buscando no nome científico...")
    pragas_selecionadas = busca_interativa_e_selecao(df, "PRAGA_NOME_CIENTIFICO", praga, 15)
    if not pragas_selecionadas:
        print("Tentando nome comum...")
        pragas_selecionadas = busca_interativa_e_selecao(df, "PRAGA_NOME_COMUM", praga, 15)
    if not pragas_selecionadas:
        return None, "Nenhuma praga selecionada"

    mask = (
        df["PRAGA_NOME_CIENTIFICO"].isin(pragas_selecionadas)
        | df["PRAGA_NOME_COMUM"].isin(pragas_selecionadas)
    )
    df_filtrado = df[mask].copy()
    resumo = (
        df_filtrado.groupby(["CULTURA", "INGREDIENTE_ATIVO"], as_index=False)
        .size()
        .rename(columns={"size": "Nº de Produtos"})
    )
    resumo = resumo.sort_values(["Nº de Produtos", "CULTURA"], ascending=[False, True])
    resumo.index += 1
    print_success(f"{len(df_filtrado):,} registros → {len(resumo):,} combinações")
    return resumo, None


def buscar_combinacoes_por_ia(df: pd.DataFrame, ia: str):
    print_header(f"MODO 3 - BUSCA POR INGREDIENTE ATIVO → {ia}")
    ias_selecionados = busca_interativa_e_selecao(df, "INGREDIENTE_ATIVO", ia, 25)
    if not ias_selecionados:
        return None, None, "Cancelado"

    print_success(f"Processando {len(ias_selecionados)} IA(s)")
    for x in ias_selecionados:
        print(f"  • {x}")

    mask_ia = df["INGREDIENTE_ATIVO"].isin(ias_selecionados)
    df_ia = df[mask_ia].copy()
    if df_ia.empty:
        return None, None, "Nenhum registro encontrado"

    grupos = df_ia.groupby(["CULTURA", "PRAGA_NOME_CIENTIFICO"])
    total_grupos = len(grupos)
    print(f"→ {total_grupos:,} combinações cultura × praga")

    combinacoes = []
    progress = smart_progress(grupos, total=total_grupos, desc="Calculando alternativas", unit="comb")
    for (cultura, praga), grupo in progress:
        mask_alt = ~df["INGREDIENTE_ATIVO"].isin(ias_selecionados)
        mask_cult = df["CULTURA"].str.strip() == cultura.strip()
        mask_praga = df["PRAGA_NOME_CIENTIFICO"].str.strip() == praga.strip()
        qtd_alt = df[mask_alt & mask_cult & mask_praga].shape[0]
        combinacoes.append({
            "CULTURA": cultura.strip(),
            "PRAGA_NOME_CIENTIFICO": praga.strip(),
            "NR_PRODUTOS_ALTERNATIVOS": qtd_alt,
        })
    if hasattr(progress, "close"):
        progress.close()

    df_combs = pd.DataFrame(combinacoes)
    df_combs = df_combs.sort_values("NR_PRODUTOS_ALTERNATIVOS", ascending=False)
    df_combs.insert(0, "Id", range(1, len(df_combs) + 1))
    df_combs = df_combs.reset_index(drop=True)
    print_success(f"Processadas {len(df_combs):,} combinações")
    return df_combs, ias_selecionados, None


def normalizar_classe(valor):
    if pd.isna(valor):
        return "não classificado"
    valor_lower = str(valor).lower().strip()
    for chave in TOX_MAP:
        if chave in valor_lower:
            return chave
    return "não classificado"


def normalizar_classe_amb(valor):
    if pd.isna(valor):
        return "não classificado"
    valor_lower = str(valor).lower().strip()
    for chave in AMB_MAP:
        if chave in valor_lower:
            return chave
    return "não classificado"


def get_tox_original(df, ia_list, cultura, praga, coluna="CLASSE_TOXICOLOGICA", is_amb=False):
    mask = (
        df["INGREDIENTE_ATIVO"].isin(ia_list)
        & (df["CULTURA"].str.lower().str.strip() == cultura.lower().strip())
        & (df["PRAGA_NOME_CIENTIFICO"].str.lower().str.strip() == praga.lower().strip())
    )
    df_orig = df[mask].copy()
    if df_orig.empty or coluna not in df_orig.columns:
        return "Desconhecida"
    norm_func = normalizar_classe_amb if is_amb else normalizar_classe
    df_orig[coluna] = df_orig[coluna].apply(norm_func)
    mode = df_orig[coluna].mode()
    return mode.iloc[0] if not mode.empty else "Desconhecida"


def adicionar_colunas_comparacao(df_alt, tox_original, amb_original):
    if tox_original == "Desconhecida":
        df_alt["Comparacao_Classe_Toxicologica"] = "Desconhecida"
    else:
        tox_num_orig = TOX_MAP.get(tox_original, 99)

        def comparar_tox(tox_alt):
            tox_norm = normalizar_classe(tox_alt)
            tox_num_alt = TOX_MAP.get(tox_norm, 99)
            if tox_num_alt == 99:
                return "Desconhecida"
            if tox_num_alt < tox_num_orig:
                return "Mais tóxico"
            if tox_num_alt > tox_num_orig:
                return "Menos tóxico"
            return "Igual"

        df_alt["Comparacao_Classe_Toxicologica"] = df_alt["CLASSE_TOXICOLOGICA"].apply(comparar_tox)

    if amb_original == "Desconhecida":
        df_alt["Comparacao_Classe_Ambiental"] = "Desconhecida"
    else:
        amb_num_orig = AMB_MAP.get(amb_original, 99)

        def comparar_amb(amb_alt):
            amb_norm = normalizar_classe_amb(amb_alt)
            amb_num_alt = AMB_MAP.get(amb_norm, 99)
            if amb_num_alt == 99:
                return "Desconhecida"
            if amb_num_alt < amb_num_orig:
                return "Mais perigoso"
            if amb_num_alt > amb_num_orig:
                return "Menos perigoso"
            return "Igual"

        df_alt["Comparacao_Classe_Ambiental"] = df_alt["CLASSE_AMBIENTAL"].apply(comparar_amb)

    cols = (
        ["Comparacao_Classe_Toxicologica", "Comparacao_Classe_Ambiental"]
        + [c for c in df_alt.columns if c not in (
            "Comparacao_Classe_Toxicologica", "Comparacao_Classe_Ambiental"
        )]
    )
    return df_alt[cols]


def gerar_csv_individual(df, ias_list, cultura, praga, pasta_base, show_msg=True, incluir_analise=False):
    mask_ia = ~df["INGREDIENTE_ATIVO"].isin(ias_list)
    mask_cult = df["CULTURA"].str.lower().str.strip() == cultura.lower().strip()
    mask_praga = df["PRAGA_NOME_CIENTIFICO"].str.lower().str.strip() == praga.lower().strip()
    df_alt = df[mask_ia & mask_cult & mask_praga].copy()
    if df_alt.empty:
        if show_msg:
            print_warning("Nenhum produto alternativo encontrado")
        return False, 0, None

    if incluir_analise:
        tox_original = get_tox_original(df, ias_list, cultura, praga, "CLASSE_TOXICOLOGICA")
        amb_original = get_tox_original(df, ias_list, cultura, praga, "CLASSE_AMBIENTAL", True)
        df_alt = adicionar_colunas_comparacao(df_alt, tox_original, amb_original)

    pasta_cultura = os.path.join(pasta_base, sanitize_filename(cultura))
    Path(pasta_cultura).mkdir(parents=True, exist_ok=True)
    nome_arq = f"Alternativos_{sanitize_filename(cultura)}_{sanitize_filename(praga)}.csv"
    caminho = os.path.join(pasta_cultura, nome_arq)
    df_alt.to_csv(caminho, sep=";", index=False, encoding="utf-8-sig")
    if show_msg:
        print_success(f"CSV gerado: {caminho} ({len(df_alt):,} linhas)")
    return True, len(df_alt), caminho


def gerar_excel_por_cultura(
    df, ias_list, cultura, df_combs_cultura, pasta_base, incluir_analise=False, ia_principal=None
):
    print_header(f"Gerando Excel: {cultura}")
    pasta_cultura = os.path.join(pasta_base, sanitize_filename(cultura))
    Path(pasta_cultura).mkdir(parents=True, exist_ok=True)

    if ia_principal and ia_principal.strip() and ia_principal.lower() != "ia":
        ia_nome = sanitize_filename(ia_principal)
    elif ias_list:
        ia_nome = sanitize_filename(ias_list[0].split("(")[0].strip())
    else:
        ia_nome = "IA_Desconhecido"

    nome_excel = f"Alternativos_{ia_nome}_{sanitize_filename(cultura)}.xlsx"
    caminho_excel = os.path.join(pasta_cultura, nome_excel)
    total_abas = len(df_combs_cultura)
    progress = smart_progress(range(total_abas), total=total_abas, desc="Criando abas", unit="praga")
    gerados = 0

    with pd.ExcelWriter(caminho_excel, engine="openpyxl") as writer:
        for idx, row in df_combs_cultura.iterrows():
            praga = row["PRAGA_NOME_CIENTIFICO"]
            mask_ia = ~df["INGREDIENTE_ATIVO"].isin(ias_list)
            mask_cult = df["CULTURA"].str.lower().str.strip() == cultura.lower().strip()
            mask_praga = df["PRAGA_NOME_CIENTIFICO"].str.lower().str.strip() == praga.lower().strip()
            df_sheet = df[mask_ia & mask_cult & mask_praga].copy()
            sheet_name = sanitize_filename(praga)[:31] or "Praga_sem_nome"

            if df_sheet.empty:
                pd.DataFrame([[f"Nenhum alternativo para {cultura} × {praga}"]], columns=["Info"]).to_excel(
                    writer, sheet_name=sheet_name, index=False
                )
            else:
                if incluir_analise:
                    tox_orig = get_tox_original(df, ias_list, cultura, praga, "CLASSE_TOXICOLOGICA")
                    amb_orig = get_tox_original(df, ias_list, cultura, praga, "CLASSE_AMBIENTAL", True)
                    df_sheet = adicionar_colunas_comparacao(df_sheet, tox_orig, amb_orig)
                df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
            gerados += 1
            progress.update(1)
        progress.close()

        if incluir_analise and gerados > 0:
            resumo_data = []
            for _, row in df_combs_cultura.iterrows():
                praga = row["PRAGA_NOME_CIENTIFICO"]
                mask = (
                    ~df["INGREDIENTE_ATIVO"].isin(ias_list)
                    & (df["CULTURA"].str.lower().str.strip() == cultura.lower().strip())
                    & (df["PRAGA_NOME_CIENTIFICO"].str.lower().str.strip() == praga.lower().strip())
                )
                df_temp = df[mask].copy()
                tox_orig = get_tox_original(df, ias_list, cultura, praga, "CLASSE_TOXICOLOGICA")
                amb_orig = get_tox_original(df, ias_list, cultura, praga, "CLASSE_AMBIENTAL", True)
                df_temp = adicionar_colunas_comparacao(df_temp, tox_orig, amb_orig)
                cont_tox = df_temp["Comparacao_Classe_Toxicologica"].value_counts().to_dict()
                cont_amb = df_temp["Comparacao_Classe_Ambiental"].value_counts().to_dict()
                resumo_data.append({
                    "Praga": praga,
                    "Total Alternativos": len(df_temp),
                    "Mais tóxico": cont_tox.get("Mais tóxico", 0),
                    "Menos tóxico": cont_tox.get("Menos tóxico", 0),
                    "Igual tóxico": cont_tox.get("Igual", 0),
                    "Desconhecida tóx": cont_tox.get("Desconhecida", 0),
                    "Mais perigoso": cont_amb.get("Mais perigoso", 0),
                    "Menos perigoso": cont_amb.get("Menos perigoso", 0),
                    "Igual amb": cont_amb.get("Igual", 0),
                    "Desconhecida amb": cont_amb.get("Desconhecida", 0),
                })
            pd.DataFrame(resumo_data).to_excel(writer, sheet_name="Resumo_Geral", index=False)

    if incluir_analise and gerados > 0:
        wb = load_workbook(caminho_excel)
        verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        amarelo = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        for sheet_name in wb.sheetnames:
            if sheet_name == "Resumo_Geral":
                continue
            ws = wb[sheet_name]
            for row in range(2, ws.max_row + 1):
                tox_cell = ws.cell(row=row, column=1)
                amb_cell = ws.cell(row=row, column=2)
                if tox_cell.value == "Menos tóxico":
                    tox_cell.fill = verde
                elif tox_cell.value == "Mais tóxico":
                    tox_cell.fill = vermelho
                elif tox_cell.value == "Igual":
                    tox_cell.fill = amarelo
                if amb_cell.value == "Menos perigoso":
                    amb_cell.fill = verde
                elif amb_cell.value == "Mais perigoso":
                    amb_cell.fill = vermelho
                elif amb_cell.value == "Igual":
                    amb_cell.fill = amarelo
        wb.save(caminho_excel)

    if gerados > 0:
        print_success(f"Excel gerado: {caminho_excel}")
    else:
        print_warning("Nenhum dado gerado para esta cultura")


# ================================================================
# NOVA FUNCIONALIDADE: PRODUTOS ALTERNATIVOS ORGANIZADOS POR PRAGA
# ================================================================

def _obter_classe_uso(serie: pd.Series) -> str:
    """Retorna a classe de uso mais frequente ou 'Não informado'."""
    if serie is None or serie.empty:
        return "Não informado"
    serie = serie.dropna().astype(str).str.strip()
    serie = serie[serie != ""]
    if serie.empty:
        return "Não informado"
    modo = serie.mode()
    return modo.iloc[0] if not modo.empty else "Não informado"


def _deduplicar_produtos_alt(df_alt: pd.DataFrame) -> pd.DataFrame:
    """Remove duplicidades de produtos alternativos, priorizando NR_REGISTRO."""
    if df_alt.empty:
        return df_alt
    if "NR_REGISTRO" in df_alt.columns:
        # Mantém a primeira ocorrência de cada registro
        return df_alt.drop_duplicates(subset=["NR_REGISTRO"], keep="first")
    subset = []
    for col in ["MARCA_COMERCIAL", "INGREDIENTE_ATIVO", "TITULAR_DE_REGISTRO"]:
        if col in df_alt.columns:
            subset.append(col)
    if subset:
        return df_alt.drop_duplicates(subset=subset, keep="first")
    return df_alt.drop_duplicates(keep="first")


# ================================================================
# COMERCIALIZAÇÃO OPCIONAL (Tabela Bruta / planilhas anuais)
# ================================================================

def _norm_col_nome(c: str) -> str:
    """Normaliza nome de coluna (remove <br>, espaços extras, acentos básicos)."""
    s = str(c).replace("<br>", " ").replace("\n", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    if UNIDECODE_AVAILABLE:
        s = unidecode(s)
    return s


def _to_float_safe(val) -> float:
    if pd.isna(val):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(".", "").replace(",", ".") if False else str(val).strip()
    # Mantém ponto decimal padrão; vírgula como decimal se não houver ponto
    if "," in s and "." not in s:
        s = s.replace(",", ".")
    s = re.sub(r"[^\d.\-]", "", s)
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


def carregar_tabela_comercializacao(
    caminhos: Optional[List[Union[str, Path]]] = None,
    pasta_busca: Optional[Union[str, Path]] = None,
) -> Optional[pd.DataFrame]:
    """
    Carrega tabela(s) de comercialização no formato 'Tabela Bruta Comercialização_Agrotóxicos AAAA'.

    Retorna DataFrame com uma linha por CHAVE_FLEXIVEL (registro MAPA normalizado):
      CHAVE_FLEXIVEL, NR_REGISTRO_COMERC, Venda_Cliente, Venda_Industria,
      Venda_Total_PF, Comercializado

    Se nenhum arquivo for encontrado, retorna None (análise segue sem vendas).
    """
    arquivos: List[Path] = []
    if caminhos:
        for c in caminhos:
            p = Path(c)
            if p.is_file():
                arquivos.append(p)
    if not arquivos:
        base = Path(pasta_busca) if pasta_busca else DADOS_DIR
        padroes = [
            "*Comercializa*.xlsx", "*Comercializa*.xls", "*Comercializa*.csv",
            "*comercializa*.xlsx", "*comercializa*.csv",
            "Tabela*Agrot*.xlsx", "Tabela*Agrot*.csv",
        ]
        for pad in padroes:
            arquivos.extend(base.glob(pad))
            arquivos.extend(Path.cwd().glob(pad))
        # Colab
        if Path("/content/dados").is_dir():
            for pad in padroes:
                arquivos.extend(Path("/content/dados").glob(pad))
        arquivos = sorted({a.resolve() for a in arquivos if a.is_file()})

    if not arquivos:
        return None

    print_header("COMERCIALIZAÇÃO (opcional)")
    dfs = []
    for arq in arquivos:
        print(f"→ Lendo: {arq.name}")
        try:
            if arq.suffix.lower() in {".xlsx", ".xls"}:
                xl = pd.ExcelFile(arq)
                # Prefere aba com 'Relatorio' / 'TODOS' / primeira
                sheet = xl.sheet_names[0]
                for s in xl.sheet_names:
                    sl = s.lower()
                    if "relatorio" in sl or "todos" in sl or "agrotox" in sl:
                        sheet = s
                        break
                df_t = pd.read_excel(arq, sheet_name=sheet, dtype=str)
            else:
                try:
                    df_t = pd.read_csv(arq, sep=";", dtype=str)
                    if len(df_t.columns) <= 1:
                        df_t = pd.read_csv(arq, sep=",", dtype=str)
                except Exception:
                    df_t = pd.read_csv(arq, sep=",", dtype=str)
        except Exception as e:
            print_error(f"Erro ao ler {arq.name}: {e}")
            continue

        # Mapeia colunas pelo nome normalizado
        col_map = {_norm_col_nome(c): c for c in df_t.columns}
        col_reg = None
        for key in ("registro mapa", "registro_mapa", "nr registro", "n registro", "registro"):
            if key in col_map:
                col_reg = col_map[key]
                break
        if col_reg is None:
            for nk, orig in col_map.items():
                if "registro" in nk and "mapa" in nk:
                    col_reg = orig
                    break
        if col_reg is None:
            print_warning(f"Coluna de registro MAPA não encontrada em {arq.name}")
            continue

        col_cli = col_map.get("vendas a cliente") or col_map.get("venda a cliente")
        col_ind = col_map.get("vendas a industria") or col_map.get("venda a industria")
        # Evita colunas ponderadas (*6, *7) preferindo as sem sufixo numérico
        if col_cli is None:
            for nk, orig in col_map.items():
                if "vendas a cliente" in nk and not re.search(r"\d$", nk.replace(" ", "")):
                    col_cli = orig
                    break
        if col_ind is None:
            for nk, orig in col_map.items():
                if "vendas a industria" in nk and not re.search(r"\d$", nk.replace(" ", "")):
                    col_ind = orig
                    break

        sub = pd.DataFrame()
        sub["NR_REGISTRO_COMERC"] = df_t[col_reg].astype(str).str.strip()
        sub["CHAVE_FLEXIVEL"] = sub["NR_REGISTRO_COMERC"].apply(extrair_chave_flexivel)
        sub["Venda_Cliente"] = df_t[col_cli].apply(_to_float_safe) if col_cli else 0.0
        sub["Venda_Industria"] = df_t[col_ind].apply(_to_float_safe) if col_ind else 0.0
        sub["Venda_Total_PF"] = sub["Venda_Cliente"] + sub["Venda_Industria"]
        # Remove linhas sem chave
        sub = sub[sub["CHAVE_FLEXIVEL"] != ""].copy()
        # Mesmo registro pode repetir por IA: agrega max (volumes PF se repetem)
        sub = (
            sub.groupby("CHAVE_FLEXIVEL", as_index=False)
            .agg({
                "NR_REGISTRO_COMERC": "first",
                "Venda_Cliente": "max",
                "Venda_Industria": "max",
                "Venda_Total_PF": "max",
            })
        )
        dfs.append(sub)
        print_success(f"  {len(sub):,} registros únicos de comercialização")

    if not dfs:
        return None

    df_v = pd.concat(dfs, ignore_index=True)
    # Se vários anos, soma volumes
    df_v = (
        df_v.groupby("CHAVE_FLEXIVEL", as_index=False)
        .agg({
            "NR_REGISTRO_COMERC": "first",
            "Venda_Cliente": "sum",
            "Venda_Industria": "sum",
            "Venda_Total_PF": "sum",
        })
    )
    df_v["Comercializado"] = df_v["Venda_Total_PF"].apply(
        lambda x: "Sim" if x and float(x) > 0 else "Não"
    )
    print_success(f"Comercialização consolidada: {len(df_v):,} produtos")
    return df_v


def _anexar_comercializacao(
    df_produtos: pd.DataFrame,
    df_vendas: Optional[pd.DataFrame],
) -> pd.DataFrame:
    """Anexa colunas de comercialização aos produtos alternativos (por NR_REGISTRO)."""
    if df_produtos is None or df_produtos.empty:
        return df_produtos

    cols_venda = ["Venda_Cliente", "Venda_Industria", "Venda_Total_PF", "Comercializado"]

    if df_vendas is None or df_vendas.empty:
        for c in cols_venda:
            if c == "Comercializado":
                df_produtos[c] = "Não analisado"
            else:
                df_produtos[c] = None
        return df_produtos

    df_out = df_produtos.copy()
    if "NR_REGISTRO" in df_out.columns:
        df_out["_CHAVE"] = df_out["NR_REGISTRO"].apply(extrair_chave_flexivel)
    else:
        df_out["_CHAVE"] = ""

    df_out = df_out.merge(
        df_vendas.rename(columns={"CHAVE_FLEXIVEL": "_CHAVE"}),
        on="_CHAVE",
        how="left",
        suffixes=("", "_comerc"),
    )
    df_out.drop(columns=["_CHAVE", "NR_REGISTRO_COMERC"], inplace=True, errors="ignore")

    for c in ("Venda_Cliente", "Venda_Industria", "Venda_Total_PF"):
        if c not in df_out.columns:
            df_out[c] = 0.0
        else:
            df_out[c] = df_out[c].fillna(0.0)
    if "Comercializado" not in df_out.columns:
        df_out["Comercializado"] = "Não"
    else:
        df_out["Comercializado"] = df_out["Comercializado"].fillna("Não")
        # Registros sem match ficam "Não" (constam na base de vendas com 0) ou
        # se não estavam na tabela: já fillna Não
    return df_out


def _metricas_comercializacao(df_prod: pd.DataFrame) -> Tuple[Any, Any]:
    """Retorna (n_produtos_comercializados, volume_total) a partir de df com colunas de venda."""
    if df_prod is None or df_prod.empty or "Comercializado" not in df_prod.columns:
        return "Não analisado", "Não analisado"
    # Se todos "Não analisado"
    if df_prod["Comercializado"].astype(str).str.lower().eq("não analisado").all():
        return "Não analisado", "Não analisado"
    n_com = int((df_prod["Comercializado"].astype(str).str.lower() == "sim").sum())
    vol = float(df_prod["Venda_Total_PF"].fillna(0).sum()) if "Venda_Total_PF" in df_prod.columns else 0.0
    return n_com, round(vol, 4)


def gerar_produtos_alternativos_por_praga(
    df: pd.DataFrame,
    ias_list: list,
    df_combs: pd.DataFrame,
    pasta_base: str,
    ia_principal: Optional[str] = None,
    incluir_analise: bool = False,
    df_vendas: Optional[pd.DataFrame] = None,
):
    """
    Gera arquivos Excel organizados por PRAGA (uma praga por arquivo).

    Cada arquivo contém:
      - Resumo_Geral_Praga : visão consolidada da praga (todas as culturas)
                             nº produtos únicos, nº IAs distintos, nº culturas
      - Resumo_IA          : Cultura | Praga | IA | Classe de USO | Grupo IRAC
                             (inclui bloco GERAL com IAs únicos da praga)
      - Relatorio          : Cultura | Praga | nº produtos alt. | nº IAs distintos |
                             Produtos comercializados | nº grupos IRAC
                             (1ª linha = GERAL com totais independentes da cultura)
      - Produtos_Alternativos : detalhamento dos produtos (auditoria)

    Colunas de comercialização e IRAC ficam preparadas para futuras integrações
    (preenchidas com 'Não analisado' / 'Não informado').

    Reutiliza df_combs já calculado e evita reprocessar combinações inexistentes.
    Não altera a geração por cultura existente.
    """
    print_header("Gerando Excel organizados por PRAGA")

    Path(pasta_base).mkdir(parents=True, exist_ok=True)

    pragas_unicas = df_combs["PRAGA_NOME_CIENTIFICO"].dropna().astype(str).str.strip().unique()
    pragas_unicas = sorted([p for p in pragas_unicas if p])

    if not pragas_unicas:
        print_warning("Nenhuma praga encontrada em df_combs.")
        return

    # Colunas preferenciais para a aba de produtos (apenas as existentes)
    colunas_preferidas = [
        "CULTURA", "PRAGA_NOME_CIENTIFICO", "PRAGA_NOME_COMUM",
        "INGREDIENTE_ATIVO", "CLASSE", "GRUPO_QUIMICO",
        "MARCA_COMERCIAL", "NR_REGISTRO", "TITULAR_DE_REGISTRO",
        "SITUACAO", "FORMULACAO", "CLASSE_TOXICOLOGICA", "CLASSE_AMBIENTAL",
    ]

    total = len(pragas_unicas)
    progress = smart_progress(range(total), total=total, desc="Pragas", unit="praga")
    gerados = 0

    for praga in pragas_unicas:
        df_praga_combs = df_combs[
            df_combs["PRAGA_NOME_CIENTIFICO"].str.strip() == praga
        ]
        culturas = (
            df_praga_combs["CULTURA"]
            .dropna()
            .astype(str)
            .str.strip()
            .unique()
            .tolist()
        )

        resumo_ia_rows: List[Dict[str, Any]] = []
        relatorio_rows: List[Dict[str, Any]] = []
        lista_produtos: List[pd.DataFrame] = []

        for cultura in culturas:
            mask_ia = ~df["INGREDIENTE_ATIVO"].isin(ias_list)
            mask_cult = df["CULTURA"].str.lower().str.strip() == cultura.lower().strip()
            mask_praga = (
                df["PRAGA_NOME_CIENTIFICO"].str.lower().str.strip()
                == praga.lower().strip()
            )
            df_alt = df[mask_ia & mask_cult & mask_praga].copy()

            if df_alt.empty:
                continue

            # Deduplicação de produtos
            df_alt_unique = _deduplicar_produtos_alt(df_alt)

            n_produtos = len(df_alt_unique)
            n_ias = df_alt_unique["INGREDIENTE_ATIVO"].nunique()

            # Comercialização por combinação (anexa só para métricas locais)
            df_alt_com_venda = _anexar_comercializacao(df_alt_unique, df_vendas)
            n_comerc, vol_comerc = _metricas_comercializacao(df_alt_com_venda)

            # --- Relatorio (uma linha por Cultura + Praga) ---
            relatorio_rows.append({
                "Cultura": cultura,
                "Praga": praga,
                "nº de produtos alternativos": n_produtos,
                "quantitativo de IA diferentes": n_ias,
                "Produtos comercializados": n_comerc,
                "Volume comercializado (PF)": vol_comerc,
                "nº grupos IRAC": "Não analisado",
            })

            # --- Resumo_IA (uma linha por Cultura + Praga + IA) ---
            # Agrupa por IA e obtém classe de uso modal
            for ia_alt, grupo in df_alt_unique.groupby("INGREDIENTE_ATIVO", sort=False):
                classe_uso = "Não informado"
                if "CLASSE" in grupo.columns:
                    classe_uso = _obter_classe_uso(grupo["CLASSE"])
                resumo_ia_rows.append({
                    "Cultura": cultura,
                    "Praga": praga,
                    "IA": ia_alt,
                    "Classe de USO": classe_uso,
                    "Grupo IRAC": "Não informado",
                })

            # Guarda produtos para a aba de detalhamento
            lista_produtos.append(df_alt_unique)

        if not relatorio_rows:
            progress.update(1)
            continue

        # Produtos_Alternativos (dedup global por registro — independente da cultura)
        if lista_produtos:
            df_produtos = pd.concat(lista_produtos, ignore_index=True)
            df_produtos = _deduplicar_produtos_alt(df_produtos)
            df_produtos = _anexar_comercializacao(df_produtos, df_vendas)
            colunas_existentes = [c for c in colunas_preferidas if c in df_produtos.columns]
            cols_venda_extra = [
                c for c in ("Comercializado", "Venda_Cliente", "Venda_Industria", "Venda_Total_PF")
                if c in df_produtos.columns
            ]
            outras = [
                c for c in df_produtos.columns
                if c not in colunas_existentes and c not in cols_venda_extra
            ]
            df_produtos = df_produtos[colunas_existentes + cols_venda_extra + outras]
        else:
            df_produtos = pd.DataFrame()

        # --- Análise GERAL da praga (independente da cultura) ---
        n_culturas = len(relatorio_rows)
        if not df_produtos.empty:
            n_produtos_geral = len(df_produtos)
            n_ias_geral = (
                df_produtos["INGREDIENTE_ATIVO"].nunique()
                if "INGREDIENTE_ATIVO" in df_produtos.columns
                else 0
            )
        else:
            n_produtos_geral = 0
            n_ias_geral = 0
        n_comerc_geral, vol_comerc_geral = _metricas_comercializacao(df_produtos)

        # Linha GERAL no topo do Relatorio
        linha_geral = {
            "Cultura": "GERAL (todas as culturas)",
            "Praga": praga,
            "nº de produtos alternativos": n_produtos_geral,
            "quantitativo de IA diferentes": n_ias_geral,
            "Produtos comercializados": n_comerc_geral,
            "Volume comercializado (PF)": vol_comerc_geral,
            "nº grupos IRAC": "Não analisado",
        }

        # Monta DataFrames finais
        df_relatorio = (
            pd.DataFrame(relatorio_rows)
            .drop_duplicates(subset=["Cultura", "Praga"])
            .sort_values("Cultura")
            .reset_index(drop=True)
        )
        # Insere a linha GERAL como primeira linha
        df_relatorio = pd.concat(
            [pd.DataFrame([linha_geral]), df_relatorio],
            ignore_index=True,
        )

        df_resumo_ia = (
            pd.DataFrame(resumo_ia_rows)
            .drop_duplicates(subset=["Cultura", "Praga", "IA"])
            .sort_values(["Cultura", "IA"])
            .reset_index(drop=True)
        )

        # Resumo_IA também ganha visão GERAL: uma linha por IA distinto da praga
        resumo_ia_geral_rows: List[Dict[str, Any]] = []
        if not df_produtos.empty and "INGREDIENTE_ATIVO" in df_produtos.columns:
            for ia_alt, grupo in df_produtos.groupby("INGREDIENTE_ATIVO", sort=False):
                classe_uso = "Não informado"
                if "CLASSE" in grupo.columns:
                    classe_uso = _obter_classe_uso(grupo["CLASSE"])
                resumo_ia_geral_rows.append({
                    "Cultura": "GERAL (todas as culturas)",
                    "Praga": praga,
                    "IA": ia_alt,
                    "Classe de USO": classe_uso,
                    "Grupo IRAC": "Não informado",
                })
            if resumo_ia_geral_rows:
                df_resumo_ia = pd.concat(
                    [
                        pd.DataFrame(resumo_ia_geral_rows).drop_duplicates(
                            subset=["Cultura", "Praga", "IA"]
                        ),
                        df_resumo_ia,
                    ],
                    ignore_index=True,
                )

        # Aba dedicada ao resumo geral da praga (visão rápida)
        obs_comerc = ""
        if n_comerc_geral != "Não analisado":
            obs_comerc = (
                f" Destes, {n_comerc_geral} produto(s) constam como comercializados "
                f"(volume PF total: {vol_comerc_geral})."
            )
        df_resumo_geral_praga = pd.DataFrame([{
            "Praga": praga,
            "Nº de culturas com alternativos": n_culturas,
            "Nº de produtos alternativos (únicos)": n_produtos_geral,
            "Nº de ingredientes ativos diferentes": n_ias_geral,
            "Produtos comercializados": n_comerc_geral,
            "Volume comercializado (PF)": vol_comerc_geral,
            "Nº grupos IRAC": "Não analisado",
            "Observação": (
                f"Para a praga {praga}, independente da cultura, "
                f"existem {n_produtos_geral} produto(s) alternativo(s) "
                f"com {n_ias_geral} ingrediente(s) ativo(s) diferente(s), "
                f"distribuídos em {n_culturas} cultura(s).{obs_comerc}"
            ),
        }])

        # Opcional: adicionar colunas de comparação toxicológica/ambiental
        if incluir_analise and not df_produtos.empty:
            dfs_com_comp = []
            for cultura in df_produtos["CULTURA"].unique():
                mask_c = df_produtos["CULTURA"] == cultura
                df_c = df_produtos[mask_c].copy()
                tox_orig = get_tox_original(
                    df, ias_list, cultura, praga, "CLASSE_TOXICOLOGICA"
                )
                amb_orig = get_tox_original(
                    df, ias_list, cultura, praga, "CLASSE_AMBIENTAL", True
                )
                df_c = adicionar_colunas_comparacao(df_c, tox_orig, amb_orig)
                dfs_com_comp.append(df_c)
            df_produtos = pd.concat(dfs_com_comp, ignore_index=True)

        # Nome do arquivo
        nome_arq = f"{sanitize_filename(praga)}.xlsx"
        caminho_excel = os.path.join(pasta_base, nome_arq)

        with pd.ExcelWriter(caminho_excel, engine="openpyxl") as writer:
            df_resumo_geral_praga.to_excel(
                writer, sheet_name="Resumo_Geral_Praga", index=False
            )
            df_resumo_ia.to_excel(writer, sheet_name="Resumo_IA", index=False)
            df_relatorio.to_excel(writer, sheet_name="Relatorio", index=False)
            if not df_produtos.empty:
                df_produtos.to_excel(
                    writer, sheet_name="Produtos_Alternativos", index=False
                )
            else:
                pd.DataFrame(
                    [["Nenhum produto alternativo encontrado"]], columns=["Info"]
                ).to_excel(writer, sheet_name="Produtos_Alternativos", index=False)

        gerados += 1
        progress.update(1)

    if hasattr(progress, "close"):
        progress.close()

    if gerados > 0:
        print_success(
            f"Gerados {gerados} arquivos Excel por praga em: {pasta_base}"
        )
    else:
        print_warning("Nenhum arquivo por praga foi gerado.")


def gerar_produtos_alternativos_por_cultura(
    df: pd.DataFrame,
    ias_list: list,
    df_combs: pd.DataFrame,
    pasta_base: str,
    ia_principal: Optional[str] = None,
    incluir_analise: bool = False,
    df_vendas: Optional[pd.DataFrame] = None,
):
    """
    Gera arquivos Excel organizados por CULTURA (mesma estrutura da análise por praga).

    Cada arquivo contém:
      - Resumo_Geral_Cultura : visão consolidada da cultura (todas as pragas)
      - Resumo_IA            : Cultura | Praga | IA | Classe de USO | Grupo IRAC
      - Relatorio            : 1ª linha GERAL + uma linha por praga
      - Produtos_Alternativos: detalhamento com comercialização (se disponível)

    Reutiliza helpers de comercialização e deduplicação compartilhados com a
    análise por praga.
    """
    print_header("Gerando Excel organizados por CULTURA (estrutura padronizada)")

    Path(pasta_base).mkdir(parents=True, exist_ok=True)

    culturas_unicas = df_combs["CULTURA"].dropna().astype(str).str.strip().unique()
    culturas_unicas = sorted([c for c in culturas_unicas if c])

    if not culturas_unicas:
        print_warning("Nenhuma cultura encontrada em df_combs.")
        return

    colunas_preferidas = [
        "CULTURA", "PRAGA_NOME_CIENTIFICO", "PRAGA_NOME_COMUM",
        "INGREDIENTE_ATIVO", "CLASSE", "GRUPO_QUIMICO",
        "MARCA_COMERCIAL", "NR_REGISTRO", "TITULAR_DE_REGISTRO",
        "SITUACAO", "FORMULACAO", "CLASSE_TOXICOLOGICA", "CLASSE_AMBIENTAL",
    ]

    total = len(culturas_unicas)
    progress = smart_progress(range(total), total=total, desc="Culturas", unit="cultura")
    gerados = 0

    for cultura in culturas_unicas:
        df_cult_combs = df_combs[df_combs["CULTURA"].str.strip() == cultura]
        pragas = (
            df_cult_combs["PRAGA_NOME_CIENTIFICO"]
            .dropna()
            .astype(str)
            .str.strip()
            .unique()
            .tolist()
        )

        resumo_ia_rows: List[Dict[str, Any]] = []
        relatorio_rows: List[Dict[str, Any]] = []
        lista_produtos: List[pd.DataFrame] = []

        for praga in pragas:
            mask_ia = ~df["INGREDIENTE_ATIVO"].isin(ias_list)
            mask_cult = df["CULTURA"].str.lower().str.strip() == cultura.lower().strip()
            mask_praga = (
                df["PRAGA_NOME_CIENTIFICO"].str.lower().str.strip()
                == praga.lower().strip()
            )
            df_alt = df[mask_ia & mask_cult & mask_praga].copy()
            if df_alt.empty:
                continue

            df_alt_unique = _deduplicar_produtos_alt(df_alt)
            n_produtos = len(df_alt_unique)
            n_ias = df_alt_unique["INGREDIENTE_ATIVO"].nunique()

            df_alt_com_venda = _anexar_comercializacao(df_alt_unique, df_vendas)
            n_comerc, vol_comerc = _metricas_comercializacao(df_alt_com_venda)

            relatorio_rows.append({
                "Cultura": cultura,
                "Praga": praga,
                "nº de produtos alternativos": n_produtos,
                "quantitativo de IA diferentes": n_ias,
                "Produtos comercializados": n_comerc,
                "Volume comercializado (PF)": vol_comerc,
                "nº grupos IRAC": "Não analisado",
            })

            for ia_alt, grupo in df_alt_unique.groupby("INGREDIENTE_ATIVO", sort=False):
                classe_uso = "Não informado"
                if "CLASSE" in grupo.columns:
                    classe_uso = _obter_classe_uso(grupo["CLASSE"])
                resumo_ia_rows.append({
                    "Cultura": cultura,
                    "Praga": praga,
                    "IA": ia_alt,
                    "Classe de USO": classe_uso,
                    "Grupo IRAC": "Não informado",
                })

            lista_produtos.append(df_alt_unique)

        if not relatorio_rows:
            progress.update(1)
            continue

        if lista_produtos:
            df_produtos = pd.concat(lista_produtos, ignore_index=True)
            df_produtos = _deduplicar_produtos_alt(df_produtos)
            df_produtos = _anexar_comercializacao(df_produtos, df_vendas)
            colunas_existentes = [c for c in colunas_preferidas if c in df_produtos.columns]
            cols_venda_extra = [
                c for c in ("Comercializado", "Venda_Cliente", "Venda_Industria", "Venda_Total_PF")
                if c in df_produtos.columns
            ]
            outras = [
                c for c in df_produtos.columns
                if c not in colunas_existentes and c not in cols_venda_extra
            ]
            df_produtos = df_produtos[colunas_existentes + cols_venda_extra + outras]
        else:
            df_produtos = pd.DataFrame()

        n_pragas = len(relatorio_rows)
        if not df_produtos.empty:
            n_produtos_geral = len(df_produtos)
            n_ias_geral = (
                df_produtos["INGREDIENTE_ATIVO"].nunique()
                if "INGREDIENTE_ATIVO" in df_produtos.columns
                else 0
            )
        else:
            n_produtos_geral = 0
            n_ias_geral = 0
        n_comerc_geral, vol_comerc_geral = _metricas_comercializacao(df_produtos)

        linha_geral = {
            "Cultura": cultura,
            "Praga": "GERAL (todas as pragas)",
            "nº de produtos alternativos": n_produtos_geral,
            "quantitativo de IA diferentes": n_ias_geral,
            "Produtos comercializados": n_comerc_geral,
            "Volume comercializado (PF)": vol_comerc_geral,
            "nº grupos IRAC": "Não analisado",
        }

        df_relatorio = (
            pd.DataFrame(relatorio_rows)
            .drop_duplicates(subset=["Cultura", "Praga"])
            .sort_values("Praga")
            .reset_index(drop=True)
        )
        df_relatorio = pd.concat(
            [pd.DataFrame([linha_geral]), df_relatorio],
            ignore_index=True,
        )

        df_resumo_ia = (
            pd.DataFrame(resumo_ia_rows)
            .drop_duplicates(subset=["Cultura", "Praga", "IA"])
            .sort_values(["Praga", "IA"])
            .reset_index(drop=True)
        )

        # Bloco GERAL de IAs da cultura
        resumo_ia_geral_rows: List[Dict[str, Any]] = []
        if not df_produtos.empty and "INGREDIENTE_ATIVO" in df_produtos.columns:
            for ia_alt, grupo in df_produtos.groupby("INGREDIENTE_ATIVO", sort=False):
                classe_uso = "Não informado"
                if "CLASSE" in grupo.columns:
                    classe_uso = _obter_classe_uso(grupo["CLASSE"])
                resumo_ia_geral_rows.append({
                    "Cultura": cultura,
                    "Praga": "GERAL (todas as pragas)",
                    "IA": ia_alt,
                    "Classe de USO": classe_uso,
                    "Grupo IRAC": "Não informado",
                })
            if resumo_ia_geral_rows:
                df_resumo_ia = pd.concat(
                    [
                        pd.DataFrame(resumo_ia_geral_rows).drop_duplicates(
                            subset=["Cultura", "Praga", "IA"]
                        ),
                        df_resumo_ia,
                    ],
                    ignore_index=True,
                )

        obs_comerc = ""
        if n_comerc_geral != "Não analisado":
            obs_comerc = (
                f" Destes, {n_comerc_geral} produto(s) constam como comercializados "
                f"(volume PF total: {vol_comerc_geral})."
            )
        df_resumo_geral = pd.DataFrame([{
            "Cultura": cultura,
            "Nº de pragas com alternativos": n_pragas,
            "Nº de produtos alternativos (únicos)": n_produtos_geral,
            "Nº de ingredientes ativos diferentes": n_ias_geral,
            "Produtos comercializados": n_comerc_geral,
            "Volume comercializado (PF)": vol_comerc_geral,
            "Nº grupos IRAC": "Não analisado",
            "Observação": (
                f"Para a cultura {cultura}, "
                f"existem {n_produtos_geral} produto(s) alternativo(s) "
                f"com {n_ias_geral} ingrediente(s) ativo(s) diferente(s), "
                f"cobrindo {n_pragas} praga(s).{obs_comerc}"
            ),
        }])

        if incluir_analise and not df_produtos.empty:
            dfs_com_comp = []
            for praga_u in df_produtos["PRAGA_NOME_CIENTIFICO"].dropna().unique():
                mask_p = df_produtos["PRAGA_NOME_CIENTIFICO"] == praga_u
                df_p = df_produtos[mask_p].copy()
                tox_orig = get_tox_original(
                    df, ias_list, cultura, praga_u, "CLASSE_TOXICOLOGICA"
                )
                amb_orig = get_tox_original(
                    df, ias_list, cultura, praga_u, "CLASSE_AMBIENTAL", True
                )
                df_p = adicionar_colunas_comparacao(df_p, tox_orig, amb_orig)
                dfs_com_comp.append(df_p)
            if dfs_com_comp:
                df_produtos = pd.concat(dfs_com_comp, ignore_index=True)

        nome_arq = f"{sanitize_filename(cultura)}.xlsx"
        caminho_excel = os.path.join(pasta_base, nome_arq)

        with pd.ExcelWriter(caminho_excel, engine="openpyxl") as writer:
            df_resumo_geral.to_excel(
                writer, sheet_name="Resumo_Geral_Cultura", index=False
            )
            df_resumo_ia.to_excel(writer, sheet_name="Resumo_IA", index=False)
            df_relatorio.to_excel(writer, sheet_name="Relatorio", index=False)
            if not df_produtos.empty:
                df_produtos.to_excel(
                    writer, sheet_name="Produtos_Alternativos", index=False
                )
            else:
                pd.DataFrame(
                    [["Nenhum produto alternativo encontrado"]], columns=["Info"]
                ).to_excel(writer, sheet_name="Produtos_Alternativos", index=False)

        gerados += 1
        progress.update(1)

    if hasattr(progress, "close"):
        progress.close()

    if gerados > 0:
        print_success(
            f"Gerados {gerados} arquivos Excel por cultura em: {pasta_base}"
        )
    else:
        print_warning("Nenhum arquivo por cultura foi gerado.")


def carregar_dados_vendas_em_memoria():
    """Compatível com modo 4: lê Comercializacao-ANO.csv legados."""
    arquivos = glob.glob("Comercializacao-*.csv") + glob.glob("Comercializacao_*.csv")
    if not arquivos:
        print_warning("Nenhum arquivo 'Comercializacao-ANO.csv' na pasta.")
        return None, []

    dfs_anos = []
    for arq in arquivos:
        match = re.search(r"(\d{4})", arq)
        if not match:
            continue
        ano = match.group(1)
        try:
            try:
                df_temp = pd.read_csv(arq, sep=";", dtype=str)
                if len(df_temp.columns) <= 1:
                    df_temp = pd.read_csv(arq, sep=",", dtype=str)
            except Exception:
                df_temp = pd.read_csv(arq, sep=",", dtype=str)

            df_temp.columns = [c.strip() for c in df_temp.columns]
            col_reg = col_venda = col_gq = None
            for c in df_temp.columns:
                c_norm = c.lower()
                if any(k in c_norm for k in ["titular", "empresa", "fabricante", "razao"]):
                    continue
                if any(k in c_norm for k in ["registro", "nr_registro", "n_registro", "num_registro"]):
                    col_reg = c
                    break
            for c in df_temp.columns:
                c_norm = c.lower()
                if any(k in c_norm for k in ["venda", "quantidade", "volume", "ton", "comercializacao"]):
                    col_venda = c
                    break
            for c in df_temp.columns:
                c_norm = c.lower()
                if "grupo" in c_norm or "classe_quimica" in c_norm:
                    col_gq = c
                    break

            if col_reg and col_venda:
                df_sub = pd.DataFrame()
                df_sub["NR_REGISTRO_ORIGINAL"] = df_temp[col_reg].astype(str).str.strip()
                df_sub["CHAVE_FLEXIVEL"] = df_sub["NR_REGISTRO_ORIGINAL"].apply(extrair_chave_flexivel)
                df_sub[f"Venda {ano}"] = df_temp[col_venda]
                if col_gq:
                    df_sub["GRUPO_QUIMICO_VENDAS"] = df_temp[col_gq].astype(str).str.strip()
                agg_dict = {"NR_REGISTRO_ORIGINAL": "first", f"Venda {ano}": "first"}
                if col_gq:
                    agg_dict["GRUPO_QUIMICO_VENDAS"] = "first"
                df_sub = (
                    df_sub[df_sub["CHAVE_FLEXIVEL"] != ""]
                    .groupby(["CHAVE_FLEXIVEL"], as_index=False)
                    .agg(agg_dict)
                )
                dfs_anos.append((ano, df_sub))
        except Exception as e:
            print_error(f"Erro ao ler {arq}: {e}")

    if not dfs_anos:
        return None, []

    dfs_anos.sort(key=lambda x: x[0])
    df_vendas = dfs_anos[0][1]
    for ano, df_ano in dfs_anos[1:]:
        df_vendas = pd.merge(
            df_vendas,
            df_ano.drop(columns=["NR_REGISTRO_ORIGINAL"], errors="ignore"),
            on="CHAVE_FLEXIVEL",
            how="outer",
        )
    cols_venda_nomes = [c for c in df_vendas.columns if c.startswith("Venda ")]
    df_vendas[cols_venda_nomes] = df_vendas[cols_venda_nomes].fillna("0")

    cols_gq = [c for c in df_vendas.columns if c.startswith("GRUPO_QUIMICO_VENDAS")]
    if cols_gq:
        df_vendas["GRUPO_QUIMICO_FINAL"] = df_vendas[cols_gq].bfill(axis=1).iloc[:, 0]
        df_vendas.drop(columns=cols_gq, inplace=True, errors="ignore")
        df_vendas.rename(columns={"GRUPO_QUIMICO_FINAL": "GRUPO_QUIMICO"}, inplace=True)

    return df_vendas, cols_venda_nomes


def gerar_produtos_com_ia(df: pd.DataFrame, ia: str, incluir_vendas: bool = False):
    print_header(f"MODO 4 - PRODUTOS QUE CONTÊM O IA → {ia}")
    ias_selecionados = busca_interativa_e_selecao(df, "INGREDIENTE_ATIVO", ia, 25)
    if not ias_selecionados:
        return None, None, "Cancelado pelo usuário"

    df_produtos = df[df["INGREDIENTE_ATIVO"].isin(ias_selecionados)].copy()
    if df_produtos.empty:
        return None, None, "Nenhum produto encontrado"

    cols_vendas = []
    if incluir_vendas:
        print("\nProcessando planilhas de comercialização legadas...")
        df_vendas, cols_vendas = carregar_dados_vendas_em_memoria()
        if df_vendas is not None:
            df_produtos["CHAVE_FLEXIVEL"] = df_produtos["NR_REGISTRO"].apply(extrair_chave_flexivel)
            df_produtos = df_produtos.merge(
                df_vendas.drop(columns=["NR_REGISTRO_ORIGINAL"], errors="ignore"),
                on="CHAVE_FLEXIVEL",
                how="left",
            )
            for c in cols_vendas:
                df_produtos[c] = df_produtos[c].fillna("0")
            df_produtos.drop(columns=["CHAVE_FLEXIVEL"], inplace=True, errors="ignore")
            print_success("Dados de venda integrados.")
        else:
            print_warning("Planilhas de vendas não encontradas.")

    if "GRUPO_QUIMICO" not in df_produtos.columns:
        col_gq = next((c for c in df.columns if "GRUPO" in c.upper() or "CLASSE_QUIMICA" in c.upper()), None)
        if col_gq and col_gq in df_produtos.columns:
            df_produtos["GRUPO_QUIMICO"] = df_produtos[col_gq]
        else:
            df_produtos["GRUPO_QUIMICO"] = "Não informado"
    else:
        df_produtos["GRUPO_QUIMICO"] = df_produtos["GRUPO_QUIMICO"].fillna("Não informado")

    colunas_base = [
        "CULTURA", "PRAGA_NOME_CIENTIFICO", "PRAGA_NOME_COMUM", "INGREDIENTE_ATIVO",
        "CLASSE", "GRUPO_QUIMICO", "MARCA_COMERCIAL", "NR_REGISTRO",
        "TITULAR_DE_REGISTRO", "SITUACAO",
    ]
    colunas_existentes = [c for c in colunas_base if c in df_produtos.columns]
    df_produtos = df_produtos[colunas_existentes + cols_vendas].drop_duplicates()
    df_produtos.sort_values(["CULTURA", "PRAGA_NOME_CIENTIFICO", "MARCA_COMERCIAL"], inplace=True)
    df_produtos.reset_index(drop=True, inplace=True)
    df_produtos.insert(0, "Id", range(1, len(df_produtos) + 1))
    for col in ["DOSE", "UNIDADE_DOSE", "OBSERVACOES"]:
        df_produtos[col] = ""
    return df_produtos, ias_selecionados, None



def main():
    # Re-resolve pastas no início (pega env vars definidas no notebook logo antes do %run)
    global SCRIPT_DIR, DADOS_DIR, SAIDA_DIR, LOGS_DIR, NO_INPUT, CSV_LOCAL
    SCRIPT_DIR, DADOS_DIR, SAIDA_DIR, LOGS_DIR, NO_INPUT, CSV_LOCAL = _resolver_pastas()

    print_header("AGROFIT — PRODUTOS ALTERNATIVOS")
    print("Script standalone para Windows / Linux / macOS / Google Colab")
    print(f"Pasta do script: {SCRIPT_DIR}")
    print(f"Pasta de dados : {DADOS_DIR}")
    print(f"Pasta de saída : {SAIDA_DIR}")
    print(f"CSV Agrofit    : {CSV_LOCAL}")
    print(f"Modo Colab     : {_detectar_colab()}")
    print(f"NO_INPUT       : {NO_INPUT}\n")

    df = carregar_dados_agrofit()
    if df is None:
        print_error("Não foi possível carregar a base Agrofit. Encerrando.")
        if not NO_INPUT:
            input("Pressione Enter para sair...")
        return

    while True:
        print_header("MENU PRINCIPAL")
        print("  1   → Busca por CULTURA")
        print("  2   → Busca por PRAGA")
        print("  3   → Busca por INGREDIENTE ATIVO + Alternativos")
        print("  4   → Produtos que CONTÊM o IA (tabela de dose)")
        print("  sair → Encerrar\n")
        print("  Nota: análises Agrofit × comercialização → analise_integrada_agrofit_comercializacao.py\n")

        modo = input(f"{bcolors.BOLD}Digite sua escolha (1/2/3/4/sair): {bcolors.ENDC}").strip().lower()

        if modo in ["sair", "s", "exit", "q", ""]:
            print_header("ENCERRADO")
            print(f"Arquivos gerados estão em: {SAIDA_DIR}")
            print("Agro é tech, agro é pop, agro é tudo! 🌱🚜")
            break

        # ---------- MODO 1 ----------
        if modo == "1":
            pasta_saida = str(SAIDA_DIR / "Resumos_Por_Cultura")
            Path(pasta_saida).mkdir(exist_ok=True)
            print_header("MODO 1 - BUSCA POR CULTURA")
            while True:
                cultura = input(f"{bcolors.OKBLUE}Nome da cultura (ou 'voltar'): {bcolors.ENDC}").strip()
                if cultura.lower() in ["voltar", "v", ""]:
                    break
                resumo, erro = gerar_resumo_por_cultura(df, cultura)
                if erro:
                    print_warning(erro)
                    continue
                nome_arq = f"Resumo_{sanitize_filename(cultura)}.csv"
                caminho = os.path.join(pasta_saida, nome_arq)
                resumo.to_csv(caminho, sep=";", index=False, encoding="utf-8-sig")
                print_success(f"Salvo: {caminho}")
                print("\nPrévia:")
                print(resumo.head(10).to_string(index=True))

        # ---------- MODO 2 ----------
        elif modo == "2":
            pasta_saida = str(SAIDA_DIR / "Resumos_Por_Praga")
            Path(pasta_saida).mkdir(exist_ok=True)
            print_header("MODO 2 - BUSCA POR PRAGA")
            while True:
                praga = input(f"{bcolors.OKBLUE}Nome da praga (ou 'voltar'): {bcolors.ENDC}").strip()
                if praga.lower() in ["voltar", "v", ""]:
                    break
                resumo, erro = gerar_resumo_por_praga(df, praga)
                if erro:
                    print_warning(erro)
                    continue
                nome_arq = f"Resumo_{sanitize_filename(praga)}.csv"
                caminho = os.path.join(pasta_saida, nome_arq)
                resumo.to_csv(caminho, sep=";", index=False, encoding="utf-8-sig")
                print_success(f"Salvo: {caminho}")
                print("\nPrévia:")
                print(resumo.head(10).to_string(index=True))

        # ---------- MODO 3 ----------
        elif modo == "3":
            pasta_base_raiz = str(SAIDA_DIR / "Produtos_alternativos_por_IA")
            Path(pasta_base_raiz).mkdir(exist_ok=True)
            print_header("MODO 3 - IA + ALTERNATIVOS")

            # Comercialização opcional (Tabela Bruta / planilhas anuais)
            print(
                "Deseja carregar tabela de comercialização? "
                "(s = buscar em dados/ / n = pular)"
            )
            if NO_INPUT:
                resp_comerc = "s"
                print_success("Modo não-interativo: tentando carregar comercialização automaticamente.")
            else:
                resp_comerc = input("> ").strip().lower()
            df_vendas_mod3 = None
            if resp_comerc in ["s", "sim", "y", "yes", ""]:
                df_vendas_mod3 = carregar_tabela_comercializacao(pasta_busca=DADOS_DIR)
                if df_vendas_mod3 is None:
                    print_warning(
                        "Nenhuma tabela de comercialização encontrada. "
                        "Análise seguirá sem dados de venda."
                    )
                else:
                    print_success(
                        f"Comercialização ativa: {len(df_vendas_mod3):,} produtos na base de vendas."
                    )

            while True:
                ia = input(f"{bcolors.OKBLUE}Ingrediente Ativo (ou 'voltar'): {bcolors.ENDC}").strip()
                if ia.lower() in ["voltar", "v", ""]:
                    break
                df_combs, ias_selecionados, erro = buscar_combinacoes_por_ia(df, ia)
                if erro:
                    print_warning(erro)
                    continue
                if df_combs is None:
                    continue

                ia_pasta = sanitize_filename(ia)
                pasta_ia = os.path.join(pasta_base_raiz, ia_pasta)

                # Por CULTURA (estrutura padronizada = mesma da por praga)
                pasta_cult_com = os.path.join(pasta_ia, "Por_Cultura", "Com_Analise")
                pasta_cult_sem = os.path.join(pasta_ia, "Por_Cultura", "Sem_Analise")
                Path(pasta_cult_com).mkdir(parents=True, exist_ok=True)
                Path(pasta_cult_sem).mkdir(parents=True, exist_ok=True)

                # Por PRAGA
                pasta_praga_com = os.path.join(pasta_ia, "Por_Praga", "Com_Analise")
                pasta_praga_sem = os.path.join(pasta_ia, "Por_Praga", "Sem_Analise")
                Path(pasta_praga_com).mkdir(parents=True, exist_ok=True)
                Path(pasta_praga_sem).mkdir(parents=True, exist_ok=True)

                # Legado (CSV individual / abas antigas) — mantido para compatibilidade
                pasta_legado_com = os.path.join(pasta_ia, "Legado_Abas_Por_Praga", "Com_Analise")
                pasta_legado_sem = os.path.join(pasta_ia, "Legado_Abas_Por_Praga", "Sem_Analise")
                Path(pasta_legado_com).mkdir(parents=True, exist_ok=True)
                Path(pasta_legado_sem).mkdir(parents=True, exist_ok=True)

                resumo_path = os.path.join(
                    pasta_ia, f"Resumo_{ia_pasta}_{len(ias_selecionados)}_selecionados.csv"
                )
                df_combs.to_csv(resumo_path, sep=";", index=False, encoding="utf-8-sig")
                print_success(f"Resumo: {resumo_path}")
                print(
                    df_combs.head(15)[
                        ["Id", "CULTURA", "PRAGA_NOME_CIENTIFICO", "NR_PRODUTOS_ALTERNATIVOS"]
                    ].to_string(index=False)
                )

                print_header("EXPORTAÇÃO")
                print("  0 / todas     → Excel POR CULTURA + POR PRAGA (estrutura padronizada)")
                print("  cultura / cu  → Excel organizados POR CULTURA (todas)")
                print("  praga / pr    → Excel organizados POR PRAGA (todas)")
                print("  c + nº        → Excel legado UMA cultura (abas por praga)")
                print("  p + nº        → CSV UMA combinação (ex: p1)")
                print("  voltar")
                while True:
                    escolha = input(
                        f"{bcolors.BOLD}Escolha: {bcolors.ENDC}"
                    ).strip().lower().replace(" ", "")
                    if escolha in ["voltar", "v", ""]:
                        break
                    if escolha in ["0", "todas", "t", "all"]:
                        gerar_produtos_alternativos_por_cultura(
                            df, ias_selecionados, df_combs, pasta_cult_com,
                            ia, True, df_vendas_mod3,
                        )
                        gerar_produtos_alternativos_por_cultura(
                            df, ias_selecionados, df_combs, pasta_cult_sem,
                            ia, False, df_vendas_mod3,
                        )
                        gerar_produtos_alternativos_por_praga(
                            df, ias_selecionados, df_combs, pasta_praga_com,
                            ia, True, df_vendas_mod3,
                        )
                        gerar_produtos_alternativos_por_praga(
                            df, ias_selecionados, df_combs, pasta_praga_sem,
                            ia, False, df_vendas_mod3,
                        )
                        print_success("Geração concluída (Por Cultura + Por Praga)!")
                    elif escolha in ["cultura", "cu", "porcultura", "por_cultura"]:
                        gerar_produtos_alternativos_por_cultura(
                            df, ias_selecionados, df_combs, pasta_cult_com,
                            ia, True, df_vendas_mod3,
                        )
                        gerar_produtos_alternativos_por_cultura(
                            df, ias_selecionados, df_combs, pasta_cult_sem,
                            ia, False, df_vendas_mod3,
                        )
                        print_success("Geração por CULTURA concluída!")
                    elif escolha in ["praga", "pr", "porpraga", "por_praga"]:
                        gerar_produtos_alternativos_por_praga(
                            df, ias_selecionados, df_combs, pasta_praga_com,
                            ia, True, df_vendas_mod3,
                        )
                        gerar_produtos_alternativos_por_praga(
                            df, ias_selecionados, df_combs, pasta_praga_sem,
                            ia, False, df_vendas_mod3,
                        )
                        print_success("Geração por PRAGA concluída!")
                    elif escolha.startswith("c") and escolha not in ("cultura", "cu"):
                        try:
                            idx = int(escolha[1:])
                            if 1 <= idx <= len(df_combs):
                                cultura = df_combs.iloc[idx - 1]["CULTURA"]
                                df_c = df_combs[df_combs["CULTURA"] == cultura]
                                gerar_excel_por_cultura(
                                    df, ias_selecionados, cultura, df_c,
                                    pasta_legado_com, True, ia,
                                )
                                gerar_excel_por_cultura(
                                    df, ias_selecionados, cultura, df_c,
                                    pasta_legado_sem, False, ia,
                                )
                                print_success(f"Excel legado: {cultura}")
                            else:
                                print_warning(f"Use 1 até {len(df_combs)}")
                        except ValueError:
                            print_warning("Formato: c1, c5")
                    elif escolha.startswith("p") and not escolha.startswith("pr"):
                        try:
                            idx = int(escolha[1:])
                            if 1 <= idx <= len(df_combs):
                                linha = df_combs.iloc[idx - 1]
                                gerar_csv_individual(
                                    df, ias_selecionados,
                                    linha["CULTURA"], linha["PRAGA_NOME_CIENTIFICO"],
                                    pasta_legado_com, True, True,
                                )
                                gerar_csv_individual(
                                    df, ias_selecionados,
                                    linha["CULTURA"], linha["PRAGA_NOME_CIENTIFICO"],
                                    pasta_legado_sem, True, False,
                                )
                            else:
                                print_warning(f"Use 1 até {len(df_combs)}")
                        except ValueError:
                            print_warning("Formato: p1, p3")
                    else:
                        print_warning("Opção inválida.")

        # ---------- MODO 4 ----------
        elif modo == "4":
            pasta_base_raiz = str(SAIDA_DIR / "Produtos_Com_IA_para_Dose")
            Path(pasta_base_raiz).mkdir(exist_ok=True)
            print_header("MODO 4 - PRODUTOS COM IA (dose)")
            while True:
                ia = input(f"{bcolors.OKBLUE}Ingrediente Ativo (ou 'voltar'): {bcolors.ENDC}").strip()
                if ia.lower() in ["voltar", "v", ""]:
                    break
                resp_vendas = input("Incluir Comercializacao-ANO.csv legados? (s/n): ").strip().lower()
                incluir_vendas = resp_vendas in ["s", "sim", "y", "yes"]
                df_produtos, ias_selecionados, erro = gerar_produtos_com_ia(df, ia, incluir_vendas)
                if erro:
                    print_warning(erro)
                    continue
                if df_produtos is None:
                    continue

                ia_pasta = sanitize_filename(ia)
                pasta_ia = os.path.join(pasta_base_raiz, ia_pasta)
                Path(pasta_ia).mkdir(exist_ok=True)
                sufixo = "Com_Vendas" if incluir_vendas else "Sem_Vendas"

                nome_csv = f"Produtos_com_{ia_pasta}_{len(ias_selecionados)}_IAs_{sufixo}.csv"
                caminho_csv = os.path.join(pasta_ia, nome_csv)
                df_produtos.to_csv(caminho_csv, sep=";", index=False, encoding="utf-8-sig")
                print_success(f"CSV: {caminho_csv}")

                nome_excel = f"Produtos_com_{ia_pasta}_por_Cultura.xlsx"
                caminho_excel = os.path.join(pasta_ia, nome_excel)
                with pd.ExcelWriter(caminho_excel, engine="openpyxl") as writer:
                    resumo = (
                        df_produtos.groupby(
                            ["CULTURA", "PRAGA_NOME_CIENTIFICO", "PRAGA_NOME_COMUM"], as_index=False
                        )
                        .agg(Qtd_Produtos=("MARCA_COMERCIAL", "nunique"), Qtd_Registros=("NR_REGISTRO", "count"))
                        .sort_values(["CULTURA", "Qtd_Produtos"], ascending=[True, False])
                    )
                    resumo.to_excel(writer, sheet_name="Resumo", index=False)
                    for cultura in sorted(df_produtos["CULTURA"].dropna().unique()):
                        df_c = df_produtos[df_produtos["CULTURA"] == cultura]
                        sheet_name = sanitize_filename(str(cultura))[:31] or "Cultura"
                        df_c.to_excel(writer, sheet_name=sheet_name, index=False)
                print_success(f"Excel: {caminho_excel}")
                print(f"Total: {len(df_produtos):,} linhas. Colunas DOSE/UNIDADE_DOSE/OBSERVACOES prontas para preenchimento.")
        else:
            print_warning("Opção inválida. Digite 1, 2, 3, 4 ou sair.")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print_warning("\nInterrompido pelo usuário.")
    except Exception as e:
        print_error(f"Erro inesperado: {e}")
        import traceback
        traceback.print_exc()
        input("\nPressione Enter para sair...")
